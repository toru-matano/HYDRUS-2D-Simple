import os
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog
import numpy as np
import pandas as pd
import openpyxl
from HYDRUS_setting import HYDRUS2DSIMPLE_INIT

class Tableview(ttk.Treeview):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.bind("<Double-1>", lambda event: self.onDoubleClick(event))

    def onDoubleClick(self, event):
        ''' Executed, when a row is double-clicked. Opens 
        read-only EntryPopup above the item's column, so it is possible
        to select text '''

        # close _Previous popups
        try:  # in case there was no _Previous popup
            self.entryPopup.destroy()
        except AttributeError:
            pass

        # what row and column was clicked on
        rowid = self.identify_row(event.y)
        column = self.identify_column(event.x)

        # handle exception when header is double click
        if not rowid:
            return

        # get column position info
        x,y,width,height = self.bbox(rowid, column)

        # y-axis offset
        pady = height // 2

        # place Entry popup properly
        text = self.item(rowid, 'values')[int(column[1:])-1]
        self.entryPopup = EntryPopup(self, rowid, int(column[1:])-1, text)
        self.entryPopup.place(x=x, y=y+pady, width=width, height=height, anchor='w')

class EntryPopup(ttk.Entry):
    def __init__(self, parent, iid, column, text, **kw):
        #ttk.Style().configure('pad.TEntry', padding='1 1 1 1')
        super().__init__(parent, style='pad.TEntry', **kw)
        self.tv = parent
        self.iid = iid
        self.column = column

        self.insert(0, text) 
        # self['state'] = 'readonly'
        # self['readonlybackground'] = 'white'
        # self['selectbackground'] = '#1BA1e2'
        self['exportselection'] = False

        self.focus_force()
        self.select_all()
        self.bind("<Return>", self.on_return)
        self.bind("<Tab>", self.on_return)
        self.bind("<Control-a>", self.select_all)
        self.bind("<Escape>", lambda *ignore: self.destroy())


    def on_return(self, event):
        rowid = self.tv.focus()
        vals = self.tv.item(rowid, 'values')
        vals = list(vals)
        vals[self.column] = self.get()
        self.tv.item(rowid, values=vals)
        self.destroy()

    def select_all(self, *ignore):
        ''' Set selection on the whole text '''
        self.selection_range(0, 'end')
        # returns 'break' to interrupt default key-bindings
        return 'break'


class HYDRUS2DSIMPLE_GUI(HYDRUS2DSIMPLE_INIT):
    def dummy(self):
        pass
    
    def Main_Menu(self):
        root = tk.Tk()
        root.title('HYDRUS 2D simple')
        root.geometry('800x500')
        
        menubar = tk.Menu(root)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="New", command=self.NewInputs)
        filemenu.add_command(label="Open", command=self.OpenInputs)
        filemenu.add_command(label="Save", command=self.SaveInputs)
        filemenu.add_command(label="Save As", command=self.SaveAsInputs)
        filemenu.add_separator()        
        filemenu.add_command(label="Exit", command=root.destroy)
        menubar.add_cascade(label="File", menu=filemenu)

        runmenu = tk.Menu(menubar, tearoff=0)
        runmenu.add_command(label="Run current project", command=self.Calculate)
        menubar.add_cascade(label="Run", menu=runmenu)
        
        editmenu = tk.Menu(menubar, tearoff=0)
        editmenu.add_command(label='Boundary Condition Options', command=self.Boundary_Condition_Options)
        menubar.add_cascade(label='Edit', menu=editmenu)
        
        def _OnDoubleClick(event):
            parents = []
            def get_parent(item):
                parent = self.trv.parent(item)
                parents.append(parent) #add parent to list
                if self.trv.parent(parent): #if parent has a parent
                    get_parent(parent)

            selected = self.trv.selection()[0]
            parents.append(selected)
            if self.trv.parent(selected):
                get_parent(selected)
            
            tree = Navigator
            while True:
                try:
                    tree = tree[parents[-1]]
                    parents = parents[:-1]
                except IndexError:
                    break
            try:
                tree()
            except TypeError:
                pass

        self.trv = ttk.Treeview(root, height=30)
        self.trv.heading('#0', text='Project Navigator - Data', anchor='w')

        Geometry = {
                'Domain Type'          : self.Domain_Type_and_Units,
                'Rectangular Domain'   : self.rectangular_domain_definition,            
            }

        Water_Flow = {
                'Iteration Criteria'   : self.iteration_criteria,
                'Soil Hydraulic Model' : self.soil_hydraulic_model,
                'Soil Hydraulic Params': self.Water_Flow_Parameters,
            }
        Solute_Transport = {
                'General Info'         : self.Solute_Transport,
                'Transport Parameters' : self.Solute_Transport_Parameters,
                'Reaction Parameters'  : self.Reqaction_Parameters_for_solute,
                'Add Fumigant'         : self.Add_Fumigant,
                'Temper. Dependence'   : self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters,
                'Wat. Cont. Dependence': self.Water_Dependent_Solute_Reaction_Parameters,
            }
        Root_Uptake = {
            'Models'                   : self.Root_Water_and_Solute_Uptake_Model,
            'Pressure Head Reduction'  : self.Root_Water_Uptake_Parameters1,
            'Pressure Head Reduction 2': self.Root_Water_Uptake_Parameters2,
            'Distribution Parameters'  : self.Root_Water_Uptake_Parameters3,
            'Root Growth Parameters'   : self.Root_Growth_Parameters,
            }

        Flow_Transport = {
                'Main process'         : self.main_process,
                'Time information'     : self.time_information,
                'Output Information'   : self.Output_Information,
                'Water Flow'           : Water_Flow,
                'Solute Transport'     : Solute_Transport,
                'Heat Transport'       : self.Heat_Transport_Parameters,
                'Root Water and Solute Uptake': Root_Uptake,
            }        
        Navigator = {
            'Project Information'          : self.New_Project,
            'Geometry'                     : Geometry,
            'Flow and Transport Parameters':Flow_Transport,
            }
        
        for key in Navigator.keys():
            self.trv.insert('', tk.END, key, text=key)
            try:
                for subkey in Navigator[key].keys():
                    self.trv.insert('', tk.END, subkey, text=subkey)
                    self.trv.move(subkey, key, tk.END)
                    try:
                        for subsubkey in Navigator[key][subkey].keys():
                            self.trv.insert('', tk.END, subsubkey, text=subsubkey)
                            self.trv.move(subsubkey, subkey, tk.END)
                    except:
                        pass
            except:
                pass
        
        self._Update_Tree()
        self.trv.bind("<Double-1>", _OnDoubleClick)
        
        self.trv.grid(row=0, column=0, padx=5, pady=5)

        root.config(menu=menubar)
        root.mainloop()
        
    def SaveNewValues(self, values):
        flag = {key:values[key] for key in values}            
        self.__dict__.update(**flag)
        
    def OpenInputs(self):
        self.path = filedialog.askdirectory()
        self.read_Dimensio()
        self.read_Meshtria()
        self.read_Boundary()
        self.read_Selector()
        self.New_Project()
        
    def NewInputs(self):
        self.__init__()
        self.New_Project()

    def SaveInputs(self, saveas=False):
        if saveas:
            self.path = filedialog.askdirectory()
        os.makedirs(self.path, exist_ok=True)
        if not os.path.isfile(self.path+r'\Domain.xlsx'):
            self.createDomainSheet()
        data = self.read_Domain_xls(self.path+r'\Domain.xlsx')
        nNode = np.where(data[0]!=0)[0]+1
        self.write_Domain(data)
        
        width = []
        for i in nNode:
            if i%self.nnode_x == 0:
                width.append((self.node_x[-1]-self.node_x[-2])/2)
            elif i%self.nnode_x == 1:
                width.append((self.node_x[1]-self.node_x[0])/2)
            else:
                width.append((self.node_x[(i+1)%self.nnode_x]-self.node_x[(i-1)%self.nnode_x])/2)
        self.write_Boundary(nNode, width)
        
        nodes, mesh = self.GenerateMesh(self.node_x, self.node_z)
        self.write_Meshtria(nodes, mesh)
        self.write_Selector(self.TPrint)
        
        if self.AtmIn:
            data = pd.read_excel(r'ATMOSPH.xlsx', 0).to_numpy()
            self.write_ATMOSPH(data)
        
        self.write_Dimensio()
        tk.messagebox.showinfo(title=None, message='Project was saved')
    
    
    def SaveAsInputs(self):
        self.SaveInputs(saveas=True)
        
        
    def Calculate(self):
        subprocess.run(['H2D_Calc64.exe', self.path], shell=True)
        print('Calculation done!')
        pass
    
    
    def _Update_Tree(self):
        if self.lWat:
            self.trv.reattach('Water Flow', 'Flow and Transport Parameters', 'end')
        else:
            self.trv.detach('Water Flow')
        if self.lChem:
            self.trv.reattach('Solute Transport', 'Flow and Transport Parameters', 'end')
            if self.lAddFum:
                self.trv.reattach('Add Fumigant', 'Solute Transport', 'end')
            else:
                self.trv.detach('Add Fumigant')                
            if self.lTDep:
                self.trv.reattach('Temper. Dependence', 'Solute Transport', 'end')
            else:
                self.trv.detach('Temper. Dependence')                
            if self.lWTDep:
                self.trv.reattach('Wat. Cont. Dependence', 'Solute Transport', 'end')
            else:
                self.trv.detach('Wat. Cont. Dependence')                
        else:
            self.trv.detach('Solute Transport')
        if self.lTemp:
            self.trv.reattach('Heat Transport', 'Flow and Transport Parameters', 'end')
        else:
            self.trv.detach('Heat Transport')
        if self.lSink:
            self.trv.reattach('Root Water and Solute Uptake', 'Flow and Transport Parameters', 'end')
            if self.rootModel:
                self.trv.reattach('Pressure Head Reduction', 'Root Water and Solute Uptake', 'end')
                self.trv.detach('Pressure Head Reduction 2')                
            else:
                self.trv.reattach('Pressure Head Reduction 2', 'Root Water and Solute Uptake', 'end')
                self.trv.detach('Pressure Head Reduction')
            if self.lMsSink:
                self.trv.reattach('Distribution Parameters', 'Root Water and Solute Uptake', 'end')
            else:                
                self.trv.detach('Distribution Parameters')
            if self.lRootGr:
                self.trv.reattach('Root Growth Parameters', 'Root Water and Solute Uptake', 'end')
            else:                
                self.trv.detach('Root Growth Parameters')            
        else:
            self.trv.detach('Root Water and Solute Uptake')

    def New_Project(self):
        root = tk.Tk()
        root.title("New Project")
        root.geometry("500x200")
        frame1 = tk.Frame(root, bd=2, relief='groove')
        frame2 = tk.Frame(root)
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)

        tk.Label(frame1, text="Project", width=40).pack()
        
        tk.Label(canvas1, text="Project Name:").pack(side='left')
        e1 = tk.Entry(canvas1, width=40, justify='right')
        e1.insert(0, self.Heading)
        e1.pack()
        
        canvas1.pack()

        tk.Label(canvas2, text="Path:").pack(side='left')
        e2 = tk.Entry(canvas2, width=40, justify='right')
        e2.insert(0, self.path)
        e2.pack(side='left')

        def _browse_button():
            # Allow user to select a directory and store it in global var
            # called folder_path
            filename = filedialog.askdirectory()
            e2.delete(0, tk.END)
            e2.insert(0, filename)
        
        button = tk.Button(canvas2, text="Browse", command=_browse_button)
        button.pack()
        
        canvas2.pack()
        
        def _GetValues():
            values = {
                'Heading':e1.get(),
                'path':e2.get()
            }
            return values
        
        def _OK():
            if e2.get() == '' or None:
                tk.messagebox.showerror(message='Select folder')
            else:
                self.SaveNewValues(_GetValues())
                root.destroy()
                self._Update_Tree()
            
        def _Next():
            if e2.get() == '' or None:
                tk.messagebox.showerror(message='Select folder')
            else:
                self.SaveNewValues(_GetValues())
                root.destroy()
                self.Domain_Type_and_Units()
        def _Cancel():
            root.destroy()

        button__Next = tk.Button(frame2, text="Next", command=_Next, width=7)
        button__OK = tk.Button(frame2, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame2, text="Cancel", command=_Cancel, width=7)
        button__Next.pack()
        button__OK.pack()
        button__Cancel.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)

        root.mainloop()

    def Domain_Type_and_Units(self):
        root = tk.Tk()
        root.title("Domain type and units")
        root.geometry("500x300")
        frame1 = tk.Frame(root, bd=2, relief='groove')
        frame2 = tk.Frame(root, bd=2, relief='groove')
        frame3 = tk.Frame(root, bd=2, relief='groove')
        frame4 = tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame3)
        canvas2 = tk.Canvas(frame3)
        canvas3 = tk.Canvas(frame3)
        
        #domain_options
        title1 = tk.Label(frame1, text="2D domain options", width=30, justify='left')
        title1.pack()
        var1  = tk.StringVar(frame1, 2)
        radio1 = tk.Radiobutton(frame1, text="horizontal plane", variable=var1, value=0, width=30, anchor='w')
        radio2 = tk.Radiobutton(frame1, text="axisymmetric vertical flow", variable=var1, value=1, width=30, anchor='w')
        radio3 = tk.Radiobutton(frame1, text="vertical plane", variable=var1, value=2, width=30, anchor='w')
        radio1.pack()
        radio2.pack()
        radio3.pack()

        #Length Unit
        title2 = tk.Label(frame2, text="Length Unit", width=30, justify='left')
        title2.pack()
        tk.Label(frame2, text="Unit:").pack(side='left')
        var2 = tk.StringVar(frame2, self.LUnit)
        Combo1 = ttk.Combobox(frame2, width=10, textvariable=var2)
        Combo1['values'] = ('mm', 'cm', 'm')
        Combo1.current()
        Combo1.pack()
        
        # Initial work space
        title3 = tk.Label(frame3, text="Initial work space", width=30, justify='left')
        title3.pack()
        tk.Label(canvas1, text="",  width=5).pack(side='left')
        tk.Label(canvas1, text="x",  width=20).pack(side='left')
        tk.Label(canvas1, text="z",  width=20, ).pack(side='left')
        
        tk.Label(canvas2, text="min:",  width=5).pack(side='left')
        e1 = tk.Entry(canvas2, width=20, justify='right')
        e1.insert(0, self.min_x)
        e1.pack(side='left')
        e2 = tk.Entry(canvas2, width=20, justify='right')
        e2.insert(0, self.min_z)
        e2.pack(side='left')

        tk.Label(canvas3, text="max:",  width=5).pack(side='left')
        e3 = tk.Entry(canvas3, width=20, justify='right')
        e3.insert(0, self.max_x)
        e3.pack(side='left')
        e4 = tk.Entry(canvas3, width=20, justify='right')
        e4.insert(0, self.max_z)
        e4.pack(side='left')
        
        canvas1.pack()
        canvas2.pack()
        canvas3.pack()
        
        def _GetValues():
            values = {
                    'Kat'   : var1.get(),
                    'LUnit' : var2.get(),
                    'min_x' : float(e1.get()),
                    'min_z' : float(e2.get()),
                    'max_x' : float(e3.get()),
                    'max_z' : float(e4.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.rectangular_domain_definition()
        
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.New_Project()
        
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=1, column=0, padx=10, pady=10)
        frame3.grid(row=2, column=0, padx=10, pady=10)
        frame4.grid(row=0, column=1, padx=10, pady=10)

        root.mainloop()
        
    def rectangular_domain_definition(self):
        root = tk.Tk()
        root.title("Rectangular domain definition")
        root.geometry("400x200")
        frame1 = tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)
        frame2 = tk.Frame(root, bd=2, relief='groove')

        tk.Label(frame1, text="Dimensions").pack()
        tk.Label(canvas1, text="x:", width=10).pack()
        e1 = tk.Entry(canvas2, width=20, justify='right')
        e1.insert(0, self.max_x-self.min_x)
        e1.pack()
        tk.Label(canvas1, text="z:", width=10).pack()
        e2 = tk.Entry(canvas2, width=20, justify='right')
        e2.insert(0, self.max_z-self.min_z)
        e2.pack()
        tk.Label(canvas1, text="α:", width=10).pack()
        e3 = tk.Entry(canvas2, width=20, justify='right')
        e3.insert(0, self.Angle)
        e3.pack()
        
        canvas1.pack(side='left')
        canvas2.pack()
        def _GetValues():
            values = {
                'Angle': e3.get(),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Domain_Discretization()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Domain_Type_and_Units()
        
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame2, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame2, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame2, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame2, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)
        root.mainloop()

    def Domain_Discretization(self):
        root = tk.Tk()
        root.title("Domain Discretization")
        root.geometry("600x400")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        frame2 =tk.Frame(root, bd=2, relief='groove')
        canvas2 = tk.Canvas(frame2)
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas3 = tk.Frame(frame1)
        canvas4 = tk.Frame(frame2)

        new_sheet = False

        tk.Label(frame1, text="Horizontal discretization in x").pack()
        tk.Label(frame2, text="Vertical discretization in z").pack()
        
        def __Update_x():
            global new_sheet
            List1.delete(0, tk.END)
            nnode_x = int(e1.get())
            node_x = np.linspace(self.min_x, self.max_x, nnode_x)
            for i in range(nnode_x):
                List1.insert(tk.END, node_x[i])
            new_sheet = True
        
        def __Update_z():
            global new_sheet
            List2.delete(0, tk.END)
            nnode_z = int(e2.get())
            node_z = np.linspace(self.min_z, self.max_z, nnode_z)
            for i in range(nnode_z):
                List2.insert(tk.END, node_z[i])
            new_sheet = True

        def _Paste_x():
            global new_sheet
            try:
                List1.delete(0, tk.END)
                e1.delete(0, tk.END)
                node_x = [float(x) for x in root.clipboard_get().split('\n')[:-1]]
                nnode_x = len(node_x)
                for i in range(nnode_x):
                    List1.insert(tk.END, node_x[i])
                e1.insert(0, nnode_x)
            except:
                tk.messagebox.showerror(message='Copy values from excel sheet')
            new_sheet = True

        def _Paste_z():
            global new_sheet
            try:
                List2.delete(0, tk.END)
                e2.delete(0, tk.END)
                node_z = [float(z) for z in root.clipboard_get().split('\n')[:-1]]
                nnode_z = len(node_z)
                for i in range(nnode_z):
                    List2.insert(tk.END, node_z[i])
                e2.insert(0, nnode_z)
            except:
                tk.messagebox.showerror(message='Copy values from excel sheet')
            new_sheet = True

        tk.Label(canvas1, text="Count:").pack(side='left')
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.nnode_x)
        e1.pack(side='left')
        button1 = tk.Button(canvas1, text="_Update", command=__Update_x)
        button1.pack(side='left')

        tk.Label(canvas2, text="Count:").pack(side='left')
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, self.nnode_z)
        e2.pack(side='left')
        button2 = tk.Button(canvas2, text="_Update", command=__Update_z)
        button2.pack(side='left')
        
        canvas1.pack()
        canvas2.pack()

        List1 = tk.Listbox(canvas3)
        List2 = tk.Listbox(canvas4)
        for i in range(self.nnode_x):
            List1.insert(tk.END, self.node_x[i])
        for i in range(self.nnode_z):
            List2.insert(tk.END, self.node_z[i])

        scrollbar1 = ttk.Scrollbar(canvas3, orient='vertical', command=List1.yview)
        scrollbar1.pack(side ='right', fill ='y')
        List1.configure(yscrollcommand = scrollbar1.set)

        scrollbar2 = ttk.Scrollbar(canvas4, orient='vertical', command=List2.yview)
        scrollbar2.pack(side ='right', fill ='y')
        List2.configure(yscrollcommand = scrollbar2.set)

        List1.pack()
        List2.pack()

        canvas3.pack()
        canvas4.pack()
        
        button3 = tk.Button(frame1, text="Paste clipboard", command=_Paste_x)
        button3.pack()
        button4 = tk.Button(frame2, text="Paste clipboard", command=_Paste_z)
        button4.pack()
        
        
        def _GetValues():
            values = {
                'node_x': List1.get(0, tk.END),
                'node_z': List2.get(0, tk.END),
                'nnode_x': int(e1.get()),
                'nnode_z': int(e2.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if new_sheet:
                self.createDomainSheet()
            self.main_process()
        
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if new_sheet:
                self.createDomainSheet()
            self.rectangular_domain_definition()
        
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)
        frame3.grid(row=0, column=2, padx=10, pady=10)

        root.mainloop()

    def createDomainSheet(self):
            
        def writeData(sheet, data):
            sheet.append(['']+list(range(1, self.nnode_x+1)))
            for i, row in enumerate(data):
                sheet.append([i+1]+row)
        
        wb = openpyxl.Workbook()
        ws_Code = wb.active
        ws_Code.title ='Code'
        Code = np.full((self.nnode_z, self.nnode_x), 0).tolist()
        writeData(ws_Code, Code)

        ws_h = wb.create_sheet('h')
        h = np.full((self.nnode_z, self.nnode_x), -100).tolist()
        writeData(ws_h, h)

        ws_Q = wb.create_sheet('Q')
        Q = np.full((self.nnode_z, self.nnode_x), 0).tolist()
        writeData(ws_Q, Q)
        
        ws_M = wb.create_sheet('M')
        M = np.full((self.nnode_z, self.nnode_x), 1).tolist()
        writeData(ws_M, M)

        ws_Temp = wb.create_sheet('Temp')
        Temp = np.full((self.nnode_z, self.nnode_x), 20).tolist()
        writeData(ws_Temp, Temp)

        ws_Conc = wb.create_sheet('Conc')
        Conc = np.full((self.nnode_z, self.nnode_x), 0).tolist()
        writeData(ws_Conc, Conc)
        
        wb.save(self.path + r'\Domain.xlsx')
        tk.messagebox.showinfo(title=None, message='New Domain.xlsx was created')

    def main_process(self):
        root = tk.Tk()
        root.title("Main Process")
        root.geometry("500x400")

        tk.Label(root, text="Simulate").grid(row=0, column=0)

        var1 = tk.IntVar(root, self.lWat)
        Check1 = tk.Checkbutton(root, text="Water Flow", variable=var1, width=30, anchor='w')
        Check1.grid(row=1, column=0, columnspan=2)

        def _Check_sol():
            if not var2.get():
                radio1.config(state=tk.DISABLED)
            #    radio2.config(state=tk.DISABLED)
            #    radio3.config(state=tk.DISABLED)
            #    radio4.config(state=tk.DISABLED)
            else:
                radio1.config(state=tk.NORMAL)
            #    radio2.config(state=tk.NORMAL)
            #    radio3.config(state=tk.NORMAL)
            #    radio4.config(state=tk.NORMAL)

        var2 = tk.IntVar(root, self.lChem)
        Check2 = tk.Checkbutton(root, text="Solute transport", variable=var2, width=30, anchor='w', command=_Check_sol)
        Check2.grid(row=2, column=0, columnspan=2)
        
        var2_1 = tk.IntVar(root, value=0)
        radio1 = tk.Radiobutton(root, text="Standard solute transport", variable=var2_1, value=0, width=30, anchor='w')
        radio2 = tk.Radiobutton(root, text="Major ion chemistry (Unsatchem)", variable=var2_1, value=1, width=30, anchor='w')
        radio3 = tk.Radiobutton(root, text="colloid-facilitated solute transport", variable=var2_1, value=2, width=30, anchor='w')
        radio4 = tk.Radiobutton(root, text="HP2", variable=var2_1, value=3, width=30, anchor='w')
        radio1.grid(row=3, column=1)
        radio2.grid(row=4, column=1)
        radio3.grid(row=5, column=1)
        radio4.grid(row=6, column=1)

        _Check_sol()
        #in development
        radio2.configure(state=tk.DISABLED)
        radio3.configure(state=tk.DISABLED)
        radio4.configure(state=tk.DISABLED)

        var3 = tk.IntVar(root, self.lTemp)
        Check3 = tk.Checkbutton(root, text="Heat transport", variable=var3, width=30, anchor='w')
        Check3.grid(row=7, column=0, columnspan=2)

        def _Check_root():
            if not var4.get():
                Check5.config(state=tk.DISABLED)
            else: Check5.config(state=tk.NORMAL)

        var4 = tk.IntVar(root, self.lSink)
        Check4 = tk.Checkbutton(root, text="Root water uptake", variable=var4, width=30, anchor='w', command=_Check_root)
        Check4.grid(row=8, column=0, columnspan=2)

        var5 = tk.IntVar(root, self.lRootGr)
        Check5 = tk.Checkbutton(root, text="Root Growth", variable=var5, width=10, anchor='w')
        Check5.grid(row=8, column=2)
        _Check_root()

        var6 = tk.IntVar(root, self.lInv)
        Check6 = tk.Checkbutton(root, text="Inverse solution", variable=var6, width=30, anchor='w')
        Check6.grid(row=9, column=0, columnspan=2)
        
        #in development
        Check6.config(state=tk.DISABLED)
        
        def _GetValues():
            values = {
                'lWat':     bool(var1.get()),
                'lChem':    bool(var2.get()),
                'lEquil':   var2_1.get()==0 and bool(var2.get()),
                'lUnsatCh': var2_1.get() == 1 and bool(var2.get()),
                'lCFSTr':   var2_1.get() == 2 and bool(var2.get()),
                'lHP2':     var2_1.get() == 3 and bool(var2.get()),
                'lTemp':    bool(var3.get()),
                'lSink':    bool(var4.get()),
                'lRootGr':  bool(var5.get()),
                'lInv':     bool(var6.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.time_information()
        
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Domain_Discretization()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(root, text="OK", command=_OK, width=7)
        button__OK.grid(row=1, column=3, padx=1, pady=1, ipadx=1)
        button_C = tk.Button(root, text="Cancel", command=_Cancel, width=7)
        button_C.grid(row=2, column=3, padx=1, pady=1, ipadx=1)
        button__Next = tk.Button(root, text="Next", command=_Next, width=7)
        button__Next.grid(row=3, column=3, padx=1, pady=1, ipadx=1)
        button_Pre = tk.Button(root, text="Previous", command=_Previous, width=7)
        button_Pre.grid(row=4, column=3, padx=1, pady=1, ipadx=1)

        root.mainloop()

    def time_information(self):
        root = tk.Tk()
        root.title("Time Information")
        root.geometry("500x400")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        frame4 =tk.Frame(root, bd=2, relief='groove')
        
        #Time Units
        tk.Label(frame1, text="Time Units").pack()
        var1 = tk.StringVar(frame1, value='Days')
        Combo1 = ttk.Combobox(frame1, width=10, textvariable=var1)
        Combo1['values'] = ('Seconds', 'Minutes', 'Hours', 'Days', 'Years')
        Combo1.current()
        Combo1.pack()

        #Time Discretization
        canvas1 = tk.Canvas(frame2)
        canvas2 = tk.Canvas(frame2)
        tk.Label(canvas1, text="Time Discretization").pack()
        tk.Label(canvas1, text="Initial time", width=20, anchor='w').pack()
        tk.Label(canvas1, text="Final time", width=20, anchor='w').pack()
        tk.Label(canvas1, text="Initial time step", width=20, anchor='w').pack()
        tk.Label(canvas1, text="Minimum time step", width=20, anchor='w').pack()
        tk.Label(canvas1, text="Maximum time step", width=20, anchor='w').pack()
        
        e1 = tk.Entry(canvas2, width=10, justify='right')
        e1.insert(0, self.tInit)
        e1.pack()
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, self.tMax)
        e2.pack()
        e3 = tk.Entry(canvas2, width=10, justify='right')
        e3.insert(0, self.dt)
        e3.pack()
        e4 = tk.Entry(canvas2, width=10, justify='right')
        e4.insert(0, self.dtMin)
        e4.pack()
        e5 = tk.Entry(canvas2, width=10, justify='right')
        e5.insert(0, self.dtMax)
        e5.pack()
        canvas1.pack(side='left')
        canvas2.pack(side='left')

        #Boundary Conditions
        def _Check_Entry():
            if not var2.get():
                e6.config(state=tk.DISABLED)
                e7.config(state=tk.DISABLED)
            else:
                e6.config(state=tk.NORMAL)
                e7.config(state=tk.NORMAL)

        tk.Label(frame3, text="Boundary Conditions").grid(row=0, column=0)
        var2 = tk.IntVar(frame3, value=self.AtmIn)
        Check1 = tk.Checkbutton(frame3, text="Time-variable boundary conditions", variable=var2, width=40, anchor='w', command=_Check_Entry)
        Check1.grid(row=1, column=0)
        Label1 = tk.Label(frame3, text="Number of time-variable boundary records", width=40, anchor='w')
        Label1.grid(row=2, column=0)
        Label2 = tk.Label(frame3, text="'Number of times to repeat the same set of BC records", width=40, anchor='w')
        Label2.grid(row=3, column=0)
        e6 = tk.Entry(frame3, width=10, justify='right')
        e6.insert(0, self.MaxAL)
        e6.grid(row=2, column=3)
        e7 = tk.Entry(frame3, width=10, justify='right')
        e7.insert(0, self.BC_Cycles)
        e7.grid(row=3, column=3)

        _Check_Entry()

        def _GetValues():
            values = {
                'TUnit':    var1.get(),
                'AtmIn':    bool(var2.get()),
                'tInit':    float(e1.get()),
                'tMax':     float(e2.get()),
                'dt':       float(e3.get()),
                'dtMin':    float(e4.get()),
                'dtMax':    float(e5.get()),
                'MaxAL':    int(e6.get()),
                'BC_Cycles':int(e7.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Output_Information()
        
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.main_process()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)
        frame3.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        frame4.grid(row=0, column=2, padx=10, pady=10)
        root.mainloop()
        
    def Output_Information(self):
        root = tk.Tk()
        root.title("Output Information")
        root.geometry("700x300")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)
        frame2 =tk.Frame(root, bd=2, relief='groove')
        canvas3 = tk.Canvas(frame2)
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas4 = tk.Canvas(frame3)
        canvas5 = tk.Canvas(frame3)
        frame4 =tk.Frame(root, bd=2, relief='groove')
        
        #Print Options
        def _Check_Entry1():
            if not var1.get():
                e1.config(state=tk.DISABLED)
            else:
                e1.config(state=tk.NORMAL)
        def _Check_Entry2():
            if not var2.get():
                e2.config(state=tk.DISABLED)
            else:
                e2.config(state=tk.NORMAL)

        tk.Label(frame1, text="Print Options",  width=30, anchor='w').pack()
        var1 = tk.IntVar(frame1, value=not self.Short)
        Check1 = tk.Checkbutton(frame1, text="T-Level information", variable=var1, width=30, anchor='w', command=_Check_Entry1)
        Check1.pack()
        
        tk.Label(canvas1, text="Every n time steps:", width=20, anchor='w').pack(side='left')
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.PrintStep)
        e1.pack()
        canvas1.pack()

        var2 = tk.IntVar(frame1, value=self.Inter)
        Check2 = tk.Checkbutton(frame1, text="Interval Output", variable=var2, width=30, anchor='w', command=_Check_Entry2)
        Check2.pack()

        tk.Label(canvas2, text="Time interval [{}]:".format(self.TUnit), width=20, anchor='w').pack(side='left')
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, self.PrintInterval)
        e2.pack()
        canvas2.pack()

        var3 = tk.IntVar(frame1, value=self.lScrn)
        Check3 = tk.Checkbutton(frame1, text="Screen Output", variable=var3, width=30, anchor='w')
        Check3.pack()
        var4 = tk.IntVar(frame1, value=self.lEnter)
        Check4 = tk.Checkbutton(frame1, text="Press Enter to end", variable=var4, width=30, anchor='w')
        Check4.pack()
        
        #Subregion for mass balance
        tk.Label(frame2, text="Subregion for mass balance",  width=30, anchor='w').pack()
        tk.Label(canvas3, text="Number of Subregions:", width=20, anchor='w').pack(side='left')
        e3 = tk.Entry(canvas3, width=10, justify='right')
        e3.insert(0, self.NLay)
        e3.pack()
        canvas3.pack()
        
        #Print times
        def _Update():
            TPrint = np.linspace(0, self.tMax, int(e4.get())+1)
            List1.delete(0, tk.END)
            for i in TPrint:
                List1.insert(tk.END, i)

        tk.Label(frame3, text="Print times",  width=30, anchor='w').pack()
        tk.Label(canvas4, text="Count:", width=5, anchor='w').pack(side='left')
        e4 = tk.Entry(canvas4, width=10, justify='right')
        e4.insert(0, self.MPL)
        e4.pack()
        button_1 = tk.Button(canvas5, text="_Update", command=_Update, width=10)
        button_1.pack()
        
        canvas4.pack(side='left', padx=5, pady=5)
        
        List1 = tk.Listbox(canvas5)
        for i in self.TPrint:
            List1.insert(tk.END, i)

        scrollbar1 = ttk.Scrollbar(canvas5, orient='vertical', command=List1.yview)
        scrollbar1.pack(side ='right', fill ='y')
        List1.configure(yscrollcommand = scrollbar1.set)
        List1.pack()
        
        canvas5.pack()
        
        def _GetValues():
            values = {
                'Short':        not var1.get(),
                'PrintStep':    int(e1.get()),
                'Inter':        bool(var2.get()),
                'PrintInterval':float(e2.get()),
                'lScrn':        bool(var3.get()),
                'lEnter':       bool(var4.get()),
                'NLay':         int(e3.get()),
                'TPrint':       List1.get(0, 'end'),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.iteration_criteria()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.time_information()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=1, column=0, padx=10, pady=10)
        frame3.grid(row=0, column=1, rowspan=2, padx=10, pady=10)
        frame4.grid(row=0, column=2, padx=10, pady=10)
        root.mainloop()
        
    def iteration_criteria(self):
        root = tk.Tk()
        root.title("Iteration criteria")
        root.geometry("400x500")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)
        frame2 =tk.Frame(root, bd=2, relief='groove')
        canvas3 = tk.Canvas(frame2)
        canvas4 = tk.Canvas(frame2)
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas5 = tk.Canvas(frame3)
        canvas6 = tk.Canvas(frame3)
        frame4 =tk.Frame(root, bd=2, relief='groove')
        frame5 =tk.Frame(root, bd=2, relief='groove')

        #Iteration criteria
        tk.Label(frame1, text="Iteration criteria",  width=30, anchor='w').pack()
        tk.Label(canvas1, text="Maximum number of iterations:", width=30, anchor='w').pack()
        e1 = tk.Entry(canvas2, width=10, justify='right')
        e1.insert(0, self.MaxIt)
        e1.pack()
        tk.Label(canvas1, text="Water content tolerance:", width=30, anchor='w').pack()
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, self.TolTh)
        e2.pack()
        tk.Label(canvas1, text="Pressure head tolerance:", width=30, anchor='w').pack()
        e3 = tk.Entry(canvas2, width=10, justify='right')
        e3.insert(0, self.TolH)
        e3.pack()
        canvas1.pack(side='left')
        canvas2.pack(side='right')
        
        #Time step control
        tk.Label(frame2, text="Time step control",  width=30, anchor='w').pack()
        tk.Label(canvas3, text="Lower optimal iteration range:", width=30, anchor='w').pack()
        e4 = tk.Entry(canvas4, width=10, justify='right')
        e4.insert(0, self.ItMin)
        e4.pack()
        tk.Label(canvas3, text="Upper optimal iteration range:", width=30, anchor='w').pack()
        e5 = tk.Entry(canvas4, width=10, justify='right')
        e5.insert(0, self.ItMax)
        e5.pack()
        tk.Label(canvas3, text="Lower timestep multiplication factor:", width=30, anchor='w').pack()
        e6 = tk.Entry(canvas4, width=10, justify='right')
        e6.insert(0, self.DMul)
        e6.pack()
        tk.Label(canvas3, text="Upper timestep multiplication factor:", width=30, anchor='w').pack()
        e7 = tk.Entry(canvas4, width=10, justify='right')
        e7.insert(0, self.DMul2)
        e7.pack()
        canvas3.pack(side='left')
        canvas4.pack(side='right')
        
        #Internal interpolation table
        tk.Label(frame3, text="Internal interpolation table",  width=30, anchor='w').pack()
        tk.Label(canvas5, text="Lower limit of the tension interval:", width=30, anchor='w').pack()
        e8 = tk.Entry(canvas6, width=10, justify='right')
        e8.insert(0, self.hTab1)
        e8.pack()
        tk.Label(canvas5, text="Upper limit of the tension interval:", width=30, anchor='w').pack()
        e9 = tk.Entry(canvas6, width=10, justify='right')
        e9.insert(0, self.hTabN)
        e9.pack()
        canvas5.pack(side='left')
        canvas6.pack(side='right')
        
        
        #Initial condition
        tk.Label(frame4, text="Initial condition",  width=30, anchor='w').pack()
        var1 = tk.IntVar(frame4, value=self.InitH_W)
        radio1 = tk.Radiobutton(frame4, text="In pressure heads", variable=var1, value=0, width=30, anchor='w')
        radio2 = tk.Radiobutton(frame4, text="In water contents", variable=var1, value=1, width=30, anchor='w')        
        radio1.pack()
        radio2.pack()

        def _GetValues():
            values = {
                'MaxIt': float(e1.get()),
                'TolTh': float(e2.get()),
                'TolH':  float(e3.get()),
                'ItMin': float(e4.get()),
                'ItMax': float(e5.get()),
                'DMul':  float(e6.get()),
                'DMul2': float(e7.get()),
                'hTab1': float(e8.get()),
                'hTabN': float(e9.get()),
                'InitH_W': bool(var1.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.soil_hydraulic_model()
            
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Output_Information()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame5, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame5, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame5, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame5, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=1, column=0, padx=10, pady=10)
        frame3.grid(row=2, column=0, padx=10, pady=10)
        frame4.grid(row=3, column=0, padx=10, pady=10)
        frame5.grid(row=0, column=1, padx=10, pady=10)
        
        root.mainloop()
        
    def soil_hydraulic_model(self):
        root = tk.Tk()
        root.title("Soil Hydraulic Model")
        root.geometry("500x600")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        
        
        #Hydraulic model
        def _Check_Model():
            if var1.get() == 0:
                Check1.config(state=tk.NORMAL)
                radio10.config(state=tk.NORMAL)
                radio11.config(state=tk.NORMAL)
                radio12.config(state=tk.NORMAL)
                radio13.config(state=tk.NORMAL)
            else:
                Check1.config(state=tk.DISABLED)
                radio10.config(state=tk.DISABLED)
                radio11.config(state=tk.DISABLED)
                radio12.config(state=tk.DISABLED)
                radio13.config(state=tk.DISABLED) 
                var2 = tk.IntVar(frame2, 0)
        
        tk.Label(frame1, text="Hydraulic model",  width=30, anchor='w').pack()
        tk.Label(frame1, text="Single porosity model",  width=30, anchor='w').pack()
        var1  = tk.IntVar(frame1, self.Model)
        radio0 = tk.Radiobutton(frame1, text="van Genuchten", variable=var1, value=0, width=30, anchor='w', command=_Check_Model)
        radio0.pack()
        
        var1_1 = tk.IntVar(frame1, self.Model==3)
        Check1 = tk.Checkbutton(frame1, text="With Air-Entry value of -2cm", variable=var1_1, width=25, anchor='w')
        Check1.pack()

        radio1 = tk.Radiobutton(frame1, text="modified van Genuchten", variable=var1, value=1, width=30, anchor='w', command=_Check_Model)
        radio2 = tk.Radiobutton(frame1, text="Brooks and Corey", variable=var1, value=2, width=30, anchor='w', command=_Check_Model)
        radio4 = tk.Radiobutton(frame1, text="Kosugi (log-normal)", variable=var1, value=4, width=30, anchor='w', command=_Check_Model)
        radio1.pack()
        radio2.pack()
        radio4.pack()
        
        tk.Label(frame1, text="Dual porosity model",  width=30, anchor='w').pack()
        radio5 = tk.Radiobutton(frame1, text="Durner, dual van Genuchten", variable=var1, value=5, width=30, anchor='w', command=_Check_Model)
        radio6 = tk.Radiobutton(frame1, text="mobile-imobile, water c, mass transfer", variable=var1, value=6, width=30, anchor='w', command=_Check_Model)
        radio7 = tk.Radiobutton(frame1, text="mobile-imobile, head mass transfer", variable=var1, value=7, width=30, anchor='w', command=_Check_Model)
        radio8 = tk.Radiobutton(frame1, text="Add-on module", variable=var1, value=8, width=30, anchor='w', command=_Check_Model)
        radio5.pack()
        radio6.pack()
        radio7.pack()
        radio8.pack()
        tk.Label(frame1, text="Other model",  width=30, anchor='w').pack()
        radio9 = tk.Radiobutton(frame1, text="Luck-up table", variable=var1, value=9, width=30, anchor='w', command=_Check_Model)
        radio9.pack()
        
        #in development
        radio9.config(state=tk.DISABLED)

        #Hysteresis
        def _Radio_Hyst():
            if var2.get() == 0:
                radio4.config(state='disabled')
                radio5.config(state='disabled')
            else:
                radio4.config(state='normal')
                radio5.config(state='normal')

        tk.Label(frame2, text="Hysteresis",  width=30, anchor='w').pack()
        var2  = tk.IntVar(frame2, 0)
        radio10 = tk.Radiobutton(frame2, text="No hysteresis", variable=var2, value=0, width=30, anchor='w', command=_Radio_Hyst)
        radio11 = tk.Radiobutton(frame2, text="Hysteresis in  retention curve", variable=var2, value=1, width=30, anchor='w', command=_Radio_Hyst)
        radio12 = tk.Radiobutton(frame2, text="Hysteresis in  retention and conductivity", variable=var2, value=2, width=30, anchor='w', command=_Radio_Hyst)
        radio13 = tk.Radiobutton(frame2, text="Bob Lenhard", variable=var2, value=3, width=30, anchor='w', command=_Radio_Hyst)
        var2_1 = tk.IntVar(frame2, -1)
        radio14 = tk.Radiobutton(frame2, text="Initially Drying Curve", variable=var2_1, value=-1, width=25, anchor='w')
        radio15 = tk.Radiobutton(frame2, text="Initially Wetting Curve", variable=var2_1, value=1, width=25, anchor='w')
        
        radio10.pack()
        radio11.pack()
        radio12.pack()
        radio13.pack()

        def _GetValues():
            values = {
                'Model':    var1.get(),
                'Hysteresis':var2.get(),
                'iKappa':   var2_1.get(),
            }
            if var1_1.get():
                values['Model'] = 3
            values['materials'] = [self.soil_catalog[str(values['Model'])]['Loam'] for _ in range(self.NMat)]
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Water_Flow_Parameters() 
            
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.iteration_criteria()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button__Cancel.pack()
        button__Next.pack()
        button_Pre.pack()
        
        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=1, column=0, padx=10, pady=10)
        frame3.grid(row=0, column=1, padx=10, pady=10)

        root.mainloop()

    def Water_Flow_Parameters(self):
        root = tk.Tk()
        root.title("Water Flow Parameters")
        root.geometry("700x350")
        materials = self.materials
        Model = str(self.Model)
        if self.Hysteresis != 0:
            Model = '8'
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        frame4 =tk.Frame(root, bd=2, relief='groove')
        
        def createTable(materials):
            for row in materials:
                trv.insert('', tk.END, values=row)

        def _Update():
            global materials
            materials = [self.soil_catalog[Model]['Loam'] for _ in range(int(e1.get()))]
            for item in trv.get_children():
                trv.delete(item)         
            createTable(materials)
            
        def ChangeMaterial(event):
            materials = [trv.item(line)['values'] for line in trv.get_children()]
            material = self.soil_catalog[Model][var1.get()]
            materials[selected_id] = material
            for item in trv.get_children():
                trv.delete(item)
            createTable(materials)
        
        tk.Label(frame1, text="Material properties for water flow",  width=80, anchor='w').pack()
        canvas1 = tk.Canvas(frame1)
        tk.Label(canvas1, text="Number of materials:",  width=20, anchor='w').pack(side=tk.LEFT)
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.NMat)
        e1.pack(side=tk.LEFT)
        button_1 = tk.Button(canvas1, text="_Update", command=_Update, width=10)
        button_1.pack()
        canvas1.pack()

        trv = Tableview(frame2, columns=tuple(range(len(materials[0]))), show='headings')
        for i, key in enumerate(self.soil_Header[Model]):
            trv.column(i, anchor='center', width=70)            
            trv.heading(i, text=key)
        
        scrollbar1 = ttk.Scrollbar(frame2, orient='vertical', command=trv.yview)
        scrollbar1.pack(side ='right', fill ='y')
        trv.configure(yscrollcommand = scrollbar1.set)

        def GetRowNum(event):
            global selected_id
            try:
                selectedItem = trv.selection()[0]
                selected_id = trv.index(selectedItem)
            except IndexError:
                pass

        trv.bind("<<TreeviewSelect>>", GetRowNum)
        trv.pack()

        createTable(materials)

        tk.Label(frame3, text="Soil Catalog",  width=10, anchor='w').pack(side=tk.LEFT)
        var1 = tk.StringVar(frame3, 'Loam')
        Combo1 = ttk.Combobox(frame3, width=10, textvariable=var1)
        Combo1['values'] = list(self.soil_catalog[Model].keys())
        Combo1.current()
        Combo1.bind("<<ComboboxSelected>>", ChangeMaterial)
        Combo1.pack(side=tk.LEFT)
        var2 = tk.IntVar(frame3, self.lWTDep)
        Check1 = tk.Checkbutton(frame3, text="Temperature dependence", variable=var2, width=25, anchor='w')
        Check1.pack()

        def _GetValues():
            values = {
                'materials': [trv.item(line)['values'] for line in trv.get_children()],
                'NMat': int(e1.get()),
                'lWTDep': bool(var2.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lChem:
                self.Solute_Transport()
            elif self.lTemp:
                self.Heat_Transport_Parameters()
            elif self.lSink:
                self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.soil_hydraulic_model()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0)
        frame2.grid(row=1, column=0)
        frame3.grid(row=2, column=0)
        frame4.grid(row=0, column=1, rowspan=3)

        root.mainloop()

    def Solute_Transport(self):
        root = tk.Tk()
        root.title("Solute Transport")
        root.geometry("600x700")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame3)
        canvas2 = tk.Canvas(frame3)
        frame4 =tk.Frame(root, bd=2, relief='groove')
        canvas3 = tk.Canvas(frame4)
        canvas4 = tk.Canvas(frame4)
        frame5 =tk.Frame(root, bd=2, relief='groove')
        frame6 =tk.Frame(root, bd=2, relief='groove')
        
        tk.Label(frame1, text="Time Weighting Scheme",  width=30, anchor='w').pack()
        var1  = tk.StringVar(frame1, self.Epsi)
        radio1 = tk.Radiobutton(frame1, text="Explicit scheme", variable=var1, value=0.0, width=30, anchor='w')
        radio2 = tk.Radiobutton(frame1, text="Crank-Nicholson scheme", variable=var1, value=0.5, width=30, anchor='w')
        radio3 = tk.Radiobutton(frame1, text="Fully implicit schem", variable=var1, value=1.0, width=30, anchor='w')
        radio1.pack()
        radio2.pack()
        radio3.pack()
        
        tk.Label(frame2, text="Space Weighting Scheme",  width=30, anchor='w').pack()
        var2 = tk.IntVar(frame2, self.lUpW)
        var3 = tk.IntVar(frame2, self.lArtD)
        radio1 = tk.Radiobutton(frame2, text="Galerkin Finite Elements", variable=var2, value=0, width=30, anchor='w')
        radio2 = tk.Radiobutton(frame2, text="Upstream weighing FE", variable=var2, value=1, width=30, anchor='w')
        radio3 = tk.Radiobutton(frame2, text="GFE with artificial Dispersion", variable=var3, value=1, width=30, anchor='w')
        radio1.pack()
        radio2.pack()
        radio3.pack()
        
        tk.Label(frame3, text="Solute Information",  width=60, anchor='w').pack()
        tk.Label(canvas1, text="Number of Solutes:",  width=20, anchor='w').pack(side='left')
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.NS)
        e1.pack(side=tk.LEFT)
        tk.Label(canvas1, text="Mass Units:",  width=20, anchor='w').pack(side='left')
        e2 = tk.Entry(canvas1, width=10, justify='right')
        e2.insert(0, self.MUnit)
        e2.pack(side=tk.LEFT)
        canvas1.pack()

        tk.Label(canvas2, text="Pulse Duration:",  width=20, anchor='w').pack(side='left')
        e3 = tk.Entry(canvas2, width=10, justify='right')
        e3.insert(0, self.tPulse)
        e3.pack(side=tk.LEFT)
        tk.Label(canvas2, text="Stability Criterion:",  width=20, anchor='w').pack(side='left')
        e4 = tk.Entry(canvas2, width=10, justify='right')
        e4.insert(0, self.PeCr)
        e4.pack(side=tk.LEFT)
        canvas2.pack()
        
        def __Check_5():
            if var7.get():
                Check5.config(state='normal')
            else:
                Check5.config(state='disabled')
        
        def _Check_7():
            if var9.get():
                Check7.config(state='normal')
            else:
                Check7.config(state='disabled')
        
        var4 = tk.IntVar(frame3, self.lTDep)
        Check1 = tk.Checkbutton(frame3, text="Temperature Dependence of Parameters", variable=var4, width=60, anchor='w')
        var5 = tk.IntVar(frame3, self.lWatDep)
        Check2 = tk.Checkbutton(frame3, text="Water Content Dependence of Parameters", variable=var5, width=60, anchor='w')
        var6 = tk.IntVar(frame3, self.Tortuosity)
        Check3 = tk.Checkbutton(frame3, text="Use Tortuosity", variable=var6, width=60, anchor='w')
        var6_1 = tk.IntVar(frame3, self.lTortM)
        radio4 = tk.Radiobutton(frame3, text="Millington & Quirk", variable=var6_1, value=0, width=50, anchor='w')
        radio5 = tk.Radiobutton(frame3, text="Moldrup", variable=var6_1, value=1, width=50, anchor='w')

        var7 = tk.IntVar(frame3, self.Bacter)        
        Check4 = tk.Checkbutton(frame3, text="Attachment/Detachment concept", variable=var7, width=60, anchor='w', command=__Check_5)
        var8 = tk.IntVar(frame3, self.Filtration)
        Check5 = tk.Checkbutton(frame3, text="Filtration Theory", variable=var8, width=50, anchor='w')

        var9 = tk.IntVar(frame3, self.lFumigant)        
        Check6 = tk.Checkbutton(frame3, text="Fumigant Module", variable=var9, width=60, anchor='w')
        var10 = tk.IntVar(frame3, 0)
        Check7 = tk.Checkbutton(frame3, text="Additional Fumigant Application at a given time", variable=var10, width=50, anchor='w')

        __Check_5()
        #in development
        Check7.config(state='disabled')
        #_Check_7()

        Check1.pack()
        Check2.pack()
        Check3.pack()
        radio4.pack()
        radio5.pack()
        Check4.pack()
        Check5.pack()
        Check6.pack()
        Check7.pack()

        tk.Label(frame4, text="Iteration Criteria for Nonlinear adsorption",  width=60, anchor='w').pack()
        tk.Label(canvas3, text="Absolute concentration tolerance:",  width=50, anchor='w').pack()
        e5 = tk.Entry(canvas4, width=10, justify='right')
        e5.insert(0, self.cTolA)
        e5.pack()
        tk.Label(canvas3, text="Relative concentration tolerance:",  width=50, anchor='w').pack()
        e6 = tk.Entry(canvas4, width=10, justify='right')
        e6.insert(0, self.cTolR)
        e6.pack()
        tk.Label(canvas3, text="Maximum number of iterations:",  width=50, anchor='w').pack()
        e7 = tk.Entry(canvas4, width=10, justify='right')
        e7.insert(0, self.MaxItC)
        e7.pack()

        canvas3.pack(side='left')
        canvas4.pack(side='right')
        
        tk.Label(frame5, text="Initial Conditions",  width=60, anchor='w').pack()
        var11  = tk.IntVar(frame5, self.lInitM)
        radio1 = tk.Radiobutton(frame5, text="In liquid phase concentrations", variable=var11, value=0, width=60, anchor='w')
        radio2 = tk.Radiobutton(frame5, text="In Total mass concentrations", variable=var11, value=1, width=60, anchor='w')
        var12 = tk.IntVar(frame5, self.lInitEq)        
        Check6 = tk.Checkbutton(frame5, text="Nonequilibrium phase is initially at equilibrium with equilibrium phase", variable=var12, width=60, anchor='w')

        radio1.pack()
        radio2.pack()
        Check6.pack()

        def _GetValues():
            values = {
                'Epsi': var1.get(),
                'lUpW': bool(var2.get()),
                'lArtD': bool(var3.get()),
                'lTDep': bool(var4.get()),
                'lWatDep': bool(var5.get()),
                'Tortuosity': bool(var6.get()),
                'lTortM': bool(var6_1.get()),
                'Bacter': bool(var7.get()),
                'Filtration': bool(var8.get()),
                'lFumigant': bool(var9.get()),
                'lInitM': bool(var11.get()),
                'lInitEq': bool(var12.get()),
                'NS': int(e1.get()),
                'MUnit': str(e2.get()), 
                'tPulse': float(e3.get()),
                'PeCr': float(e4.get()), 
                'cTolA': float(e5.get()),
                'cTolR': float(e6.get()),
                'MaxItC': float(e7.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Solute_Transport_Parameters()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Water_Flow_Parameters()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()
            
        button__OK = tk.Button(frame6, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame6, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame6, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame6, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)
        frame3.grid(row=1, column=0, columnspan=2, padx=10, pady=10)
        frame4.grid(row=2, column=0, columnspan=2, padx=10, pady=10)
        frame5.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
        frame6.grid(row=0, column=2, padx=10, pady=10)

        root.mainloop() 
        
    def Solute_Transport_Parameters(self):
        
        root = tk.Tk()
        root.title("Solute Transport Parameters")
        root.geometry("900x300")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')

        header_soil = ['Bulk Density', 'Dispersivity Length', 'Dispersivity T', 'Fract', 'ThImob.']
        header_solute = ['Diffus.W.', 'Diffus.G.']

        tk.Label(frame1, text="Soil specific parameters", width=70, anchor='w').pack()
        trv1 = Tableview(frame1, columns=tuple(range(len(header_soil))), show='headings')
        for i, key in enumerate(header_soil):
            trv1.column(i, anchor='center', width=100)            
            trv1.heading(i, text=key)
        
        for row in self.soil_param:
            trv1.insert('', tk.END, values=row)

        tk.Label(frame2, text="Solute specific parameters", width=25, anchor='w').pack()
        trv2 = Tableview(frame2, columns=tuple(range(len(header_solute))), show='headings')
        for i, key in enumerate(header_solute):
            trv2.column(i, anchor='center', width=100)            
            trv2.heading(i, text=key)
        
        for row in self.solute_param:
            trv2.insert('', tk.END, values=row)

        trv1.pack()
        trv2.pack()

        def _GetValues():
            values = {
                'soil_param':  [trv1.item(line)['values'] for line in trv1.get_children()],
                'solute_param':  [trv2.item(line)['values'] for line in trv2.get_children()],
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Reqaction_Parameters_for_solute()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Solute_Transport()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()        
            self._Update_Tree()

        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=0, column=1, padx=10, pady=10)
        frame3.grid(row=0, column=2, padx=10, pady=10)

        root.mainloop() 
        
    def Reqaction_Parameters_for_solute(self):
        root = tk.Tk()
        root.title("Reaction parameters for solute")
        root.geometry("650x350")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        frame4 =tk.Frame(root, bd=2, relief='groove')

        header_BC = ['cTop', 'cBot', 'cBnd3', 'cBnd4', 'cRoot', 'cWell', 'cBnd7', 'cAtm', 'd']
        header_ReactPrm = ['Ks', 'Nu', 'β', 'Henry', 'SnkL1', 'SnkS1', 'SnkG1', 'SnkL1\'', 'SnkS1\'', 'SnkG1\'', 'SnkL0', 'SnkS0', 'SnkG0', 'α']
        if self.Bacter:
            if self.Filtration:
                header_ReactPrm = ['Ks', 'Nu', 'β', 'Henry', 'SnkL1', 'SnkS1', 'D_soil', 'D_virus', 'SMax2', 'Coll.Eff2', 'DetachS2', 'SMax1', 'Coll.Eff1', 'DetachS1']
            else:
                header_ReactPrm = ['Ks', 'Nu', 'β', 'Henry', 'SnkL1', 'SnkS1', 'iPsi2', 'iPsi1', 'SMax2', 'AttachS2', 'DetachS2', 'SMax1', 'AttachS1', 'DetachS1']

        tk.Label(frame1, text='Select solute: ').pack(side='left')
        var1 = tk.IntVar(frame1, 1)
        Combo1 = ttk.Combobox(frame1, width=10, textvariable=var1)
        Combo1['values'] = tuple(range(1, self.NS+1))
        Combo1.current()
        Combo1.pack()

        tk.Label(frame2, text="Boundary Conditions", width=70, anchor='w').pack()
        trv1 = Tableview(frame2, columns=tuple(range(len(header_BC))), show='headings', height=1)
        for i, key in enumerate(header_BC):
            trv1.column(i, anchor='center', width=50)            
            trv1.heading(i, text=key)
        
        for row in self.cBnds:
            trv1.insert('', tk.END, values=row)

        tk.Label(frame3, text="Reaction parameters", width=70, anchor='w').pack()
        trv2 = Tableview(frame3, columns=tuple(range(len(header_ReactPrm))), show='headings', height=self.NMat)
        for i, key in enumerate(header_ReactPrm):
            trv2.column(i, anchor='center', width=50)            
            trv2.heading(i, text=key)
        
        for row in self.react_param:
            trv2.insert('', tk.END, values=row)

        scrlbar1 = ttk.Scrollbar(frame2, orient ="horizontal", command = trv1.xview)
        scrlbar1.pack(side ='bottom', fill ='x')
        trv1.configure(xscrollcommand = scrlbar1.set)
        scrlbar2 = ttk.Scrollbar(frame3, orient ="horizontal", command = trv2.xview)
        scrlbar2.pack(side ='bottom', fill ='x')
        trv2.configure(xscrollcommand = scrlbar2.set)

        trv1.place(x=0, y=20, width=500, height=100)
        trv2.place(x=0, y=20, width=500, height=100)
        
        def _Check_iPsi():
            if self.Bacter and not self.Filtration:
                #if iPsi >5
                if trv1.item(trv1.get_children()[13])['values'] > 5:
                    ttk.messagebox.showerror(message='Value must be in 0, 1, 2, 3 or 4')
                    return False
            return True
            
        def _GetValues():
            values = {
                'cBnds':  [trv1.item(line)['values'] for line in trv1.get_children()],
                'react_param':  [trv2.item(line)['values'] for line in trv2.get_children()],
            }
            return values
        
            
        def _Next():
            if _Check_iPsi():
                self.SaveNewValues(_GetValues())
                root.destroy()
                
                if self.lAddFum:
                    self.Add_Fumigant()
                elif self.lTDep:
                    self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters()
                elif self.lWatDep:
                    self.Water_Dependent_Solute_Reaction_Parameters()
                elif self.lTemp:
                    self.Heat_Transport_Parameters()
                elif self.lSink:
                    self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            if _Check_iPsi():
                self.SaveNewValues(_GetValues())
                root.destroy()
                self.Solute_Transport_Parameters()
            
        def _OK():
            if _Check_iPsi():
                self.SaveNewValues(_GetValues())
                root.destroy()
                self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.place(x=20, y=20, width=300, height=20)
        frame2.place(x=20, y=50, width=500, height=110)
        frame3.place(x=20, y=180, width=500, height=110)
        frame4.place(x=550, y=30, height=300)

        root.mainloop()
    
    
    def Add_Fumigant(self):
        root = tk.Tk()
        root.title("Add Fumigant")
        root.geometry("500x250")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        
        tk.Label(frame1, text='Additional Fumigant Application', anchor='w', width=50).grid(row=0, column=0, columnspan=2)
        tk.Label(frame1, text='Time of Additional Fumigant Application:', anchor='w', width=40).grid(row=1, column=0)
        tk.Label(frame1, text='Mass of Applied Fumigant:', anchor='w', width=40).grid(row=2, column=0)
        e1 = tk.Entry(frame1, justify='right', width=10)
        e1.insert(0, self.AddFumTime)
        e1.grid(row=1, column=1)
        e2 = tk.Entry(frame1, justify='right', width=10)
        e2.insert(0, self.AddFumMass)
        e2.grid(row=2, column=1)        
        
        tk.Label(frame2, text='Location of Fumigant Application', anchor='w', width=50).grid(row=0, column=0, columnspan=2)
        tk.Label(frame2, text='Min x-coordinate of application:', anchor='w', width=40).grid(row=1, column=0)
        tk.Label(frame2, text='Max x-coordinate of application:', anchor='w', width=40).grid(row=2, column=0)
        tk.Label(frame2, text='Min z-coordinate of application:', anchor='w', width=40).grid(row=3, column=0)
        tk.Label(frame2, text='Max z-coordinate of application:', anchor='w', width=40).grid(row=4, column=0)
        e3 = tk.Entry(frame2, justify='right', width=10)
        e3.insert(0, self.AddFumMinX)
        e3.grid(row=1, column=1)
        e4 = tk.Entry(frame2, justify='right', width=10)
        e4.insert(0, self.AddFumMaxX)
        e4.grid(row=2, column=1)
        e5 = tk.Entry(frame2, justify='right', width=10)
        e5.insert(0, self.AddFumMinZ)
        e5.grid(row=3, column=1)
        e6 = tk.Entry(frame2, justify='right', width=10)
        e6.insert(0, self.AddFumMaxZ)
        e6.grid(row=4, column=1)

        def _GetValues():
            values = {
                'AddFumTime': float(e1.get()),
                'AddFumMass': float(e2.get()),
                'AddFumMinX': float(e3.get()),
                'AddFumMaxX': float(e4.get()),
                'AddFumMinZ': float(e5.get()),
                'AddFumMaxZ': float(e6.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
        
            if self.lTDep:
                self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters()
            elif self.lWatDep:
                self.Water_Dependent_Solute_Reaction_Parameters()
            elif self.lTemp:
                self.Heat_Transport_Parameters()
            elif self.lSink:
                self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Reqaction_Parameters_for_solute()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=20, pady=10)
        frame2.grid(row=1, column=0, padx=20, pady=10)
        frame3.grid(row=0, column=1, rowspan=2, padx=10, pady=10)

        root.mainloop()        
        
    def Temperature_Dependent_Solute_Transport_and_Reaction_Parameters(self):
        root = tk.Tk()
        root.title("Temperature Dependent Solute Transport and Reaction Parameters (in J/mol)")
        root.geometry("700x300")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        header = ['Diff.W.', 'Diff.G.', 'Ks', 'Nu', 'β', 'Henry', 'SnkL1', 'SnkS1', 'SnkG1', 'SnkL1\'', 'SnkS1\'', 'SnkG1\'', 'SnkL0', 'SnkS0', 'SnkG0', 'α']

        tk.Label(frame1, text="Parameters", width=70, anchor='w').pack()
        trv = Tableview(frame1, columns=tuple(range(len(header))), show='headings', height=self.NMat)
        for i, key in enumerate(header):
            trv.column(i, anchor='center', width=50)            
            trv.heading(i, text=key)
        
        for row in self.TDep_params:
            trv.insert('', tk.END, values=row)

        scrlbar1 = ttk.Scrollbar(frame1, orient ="horizontal", command = trv.xview)
        scrlbar1.pack(side ='bottom', fill ='x')
        trv.configure(xscrollcommand = scrlbar1.set)
        trv.place(x=10, y=20, width=500, height=500)

        def _GetValues():
            values = {
                'TDep_params':  [trv.item(line)['values'] for line in trv.get_children()],
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            
            if self.lWatDep:
                self.Water_Dependent_Solute_Reaction_Parameters()
            elif self.lTemp:
                self.Heat_Transport_Parameters()
            elif self.lSink:
                self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lAddFum:
                    self.Add_Fumigant()
            else:
                self.Reqaction_Parameters_for_solute()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame2, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame2, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame2, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame2, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.place(x=10, y=10, width=520, height=200)
        frame2.place(x=550, y=30, height=300)

        root.mainloop()

    def Water_Dependent_Solute_Reaction_Parameters(self):
        root = tk.Tk()
        root.title("Water Content Dependent Solute Reaction Parameters")
        root.geometry("650x350")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')

        header = ['SnkL1', 'SnkS1', 'SnkG1', 'SnkL1\'', 'SnkS1\'', 'SnkG1\'', 'SnkL0', 'SnkS0', 'SnkG0']

        tk.Label(frame1, text="Parameter 1 (β)", width=70, anchor='w').pack()
        trv1 = Tableview(frame1, columns=tuple(range(len(header))), show='headings', height=self.NS)
        for i, key in enumerate(header):
            trv1.column(i, anchor='center', width=50)            
            trv1.heading(i, text=key)
        
        for row in self.WTDep_params1:
            trv1.insert('', tk.END, values=row)

        tk.Label(frame2, text="Parameter 2 (h_ref)", width=70, anchor='w').pack()
        trv2 = Tableview(frame2, columns=tuple(range(len(header))), show='headings', height=self.NS)
        for i, key in enumerate(header):
            trv2.column(i, anchor='center', width=50)            
            trv2.heading(i, text=key)
        
        for row in self.WTDep_params2:
            trv2.insert('', tk.END, values=row)

        scrlbar1 = ttk.Scrollbar(frame1, orient ="horizontal", command = trv1.xview)
        scrlbar1.pack(side ='bottom', fill ='x')
        trv1.configure(xscrollcommand = scrlbar1.set)
        scrlbar2 = ttk.Scrollbar(frame2, orient ="horizontal", command = trv2.xview)
        scrlbar2.pack(side ='bottom', fill ='x')
        trv2.configure(xscrollcommand = scrlbar2.set)

        trv1.place(x=0, y=20, width=500, height=100)
        trv2.place(x=0, y=20, width=500, height=100)

        def _GetValues():
            values = {
                'WTDep_params1':  [trv1.item(line)['values'] for line in trv1.get_children()],
                'WTDep_params2':  [trv2.item(line)['values'] for line in trv2.get_children()],
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            
            if self.lTemp:
                self.Heat_Transport_Parameters()
            elif self.lSink:
                self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lTDep:
                self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters()
            elif self.lAddFum:
                self.Add_Fumigant()
            else:
                self.Reqaction_Parameters_for_solute()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.place(x=20, y=20, width=500, height=110)
        frame2.place(x=20, y=150, width=500, height=110)
        frame3.place(x=550, y=30, height=300)

        root.mainloop()

    def Heat_Transport_Parameters(self):
        root = tk.Tk()
        root.title("Heat transport Parameters")
        root.geometry("650x400")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame2)
        canvas2 = tk.Canvas(frame2)
        frame3 =tk.Frame(root, bd=2, relief='groove')
        frame4 =tk.Frame(root, bd=2, relief='groove')

        header = ['tBnd1', 'tBnd2', 'tBnd3', 'tBnd4', 'tBnd5', 'tWell']
        header2 = ['Qn', 'Qo', 'Disper.', 'B1', 'B2', 'B3', 'Cn', 'Co', 'Cw']

        def ChangeMaterial(event):
            try:
                Tpars = [trv2.item(line)['values'] for line in trv2.get_children()]
                Tpar = self.Heat_Catalog[var1.get()]
                Tpars[selected_id] = Tpar
                for item in trv2.get_children():
                    trv2.delete(item)
                for row in Tpars:
                    trv2.insert('', tk.END, values=row)
            except NameError:
                pass

        tk.Label(frame1, text="Boundary Conditions", width=70, anchor='w').pack()
        trv1 = Tableview(frame1, columns=tuple(range(len(header))), show='headings', height=1)
        for i, key in enumerate(header):
            trv1.column(i, anchor='center', width=50)            
            trv1.heading(i, text=key)
        
        for row in self.tBnds:
            trv1.insert('', tk.END, values=row)

        tk.Label(frame2, text='Heat transport parameters 1', width=70, anchor='w').pack()
        
        tk.Label(canvas1, text='Temperature Amplitude [C]:', width=40, anchor='w').pack()
        tk.Label(canvas1, text='Time interval for one temp. cycle [day]', width=40, anchor='w').pack()
        tk.Label(canvas1, text='Thermal conductivity: ', width=40, anchor='w').pack()

        e1 = tk.Entry(canvas2, width=12, justify='right')
        e1.insert(0, self.tAmpl)
        e1.pack()
        e2 = tk.Entry(canvas2, width=12, justify='right')
        e2.insert(0, self.tPeriodHeat)
        e2.pack()
        var1 = tk.StringVar(canvas2, 'loam')
        Combo1 = ttk.Combobox(canvas2, width=10, textvariable=var1)
        Combo1['values'] =  list(self.Heat_Catalog.keys())
        Combo1.current()
        Combo1.bind("<<ComboboxSelected>>", ChangeMaterial)
        Combo1.pack()
        
        canvas1.pack(side='left')
        canvas2.pack(side='right')
        
        tk.Label(frame3, text="Heat transport parameters 2", width=70, anchor='w').pack()
        trv2 = Tableview(frame3, columns=tuple(range(len(header2))), show='headings', height=self.NMat)
        for i, key in enumerate(header2):
            trv2.column(i, anchor='center', width=50)            
            trv2.heading(i, text=key)
        
        for row in self.Tpar:
            trv2.insert('', tk.END, values=row)

        scrlbar1 = ttk.Scrollbar(frame1, orient ="horizontal", command = trv1.xview)
        scrlbar1.pack(side ='bottom', fill ='x')
        trv1.configure(xscrollcommand = scrlbar1.set)
        scrlbar2 = ttk.Scrollbar(frame3, orient ="horizontal", command = trv2.xview)
        scrlbar2.pack(side ='bottom', fill ='x')
        trv2.configure(xscrollcommand = scrlbar2.set)

        def GetRowNum(event):
            global selected_id
            try:
                selectedItem = trv2.selection()[0]
                selected_id = trv2.index(selectedItem)
            except IndexError:
                pass

        trv2.bind("<<TreeviewSelect>>", GetRowNum)

        trv1.place(x=0, y=20, width=500, height=80)
        trv2.place(x=0, y=20, width=500, height=150)

        def _GetValues():
            values = {
                'tAmpl': float(e1.get()),
                'tPeriodHeat': float(e2.get()),
                'tBnds':  [trv1.item(line)['values'] for line in trv1.get_children()],
                'Tpar':  [trv2.item(line)['values'] for line in trv2.get_children()],
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lSink:
                self.Root_Water_and_Solute_Uptake_Model()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lWTDep:
                self.Water_Dependent_Solute_Reaction_Parameters()
            elif self.lTDep:
                self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters()
            elif self.lAddFum:
                self.Add_Fumigant()
            elif self.lChem:
                self.Reqaction_Parameters_for_solute()
            elif self.lWat:
                self.Water_Flow_Parameters()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.place(x=20, y=20, width=500, height=90)
        frame2.place(x=20, y=110, width=500, height=120)
        frame3.place(x=20, y=220, width=500, height=130)
        frame4.place(x=550, y=30, height=300)

        root.mainloop()

    def Root_Water_and_Solute_Uptake_Model(self):
        root = tk.Tk()
        root.title("Root Water and Solute Uptake Model")
        root.geometry("500x550")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas2 = tk.Canvas(frame3)
        frame4 =tk.Frame(root, bd=2, relief='groove')
        
        tk.Label(frame1, text="Water Uptake Reduction Model", width=50, anchor='w').pack(pady=10)
        var1  = tk.IntVar(frame1, self.rootModel)
        radio1 = tk.Radiobutton(canvas1, text="Feddes", variable=var1, value=0, width=30, anchor='w')
        radio2 = tk.Radiobutton(canvas1, text="S-Shaped", variable=var1, value=1, width=30, anchor='w')
        radio1.grid(row=1, column=0)
        radio2.grid(row=2, column=0)
        tk.Label(canvas1, text="Critical Stress Index:", width=30, anchor='w').grid(row=3, column=0)
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.OmegaC)
        e1.grid(row=3, column=1)
        
        canvas1.pack()

        tk.Label(frame2, text="Solute Stress Model", width=50, anchor='w').pack(pady=10)
        if not self.SoluteReduction:
            stress_model = 0
        else:
            if self.SoluteAdditive:
                stress_model = 1
            else:
                stress_model = 2
        var2  = tk.IntVar(frame2, stress_model)
        radio1 = tk.Radiobutton(frame2, text="No Solute Stress", variable=var2, value=0, width=40, anchor='w')
        radio2 = tk.Radiobutton(frame2, text="Additive Model", variable=var2, value=1, width=40, anchor='w')
        radio3 = tk.Radiobutton(frame2, text="Multiplicated Model", variable=var2, value=2, width=40, anchor='w')
        var3 = tk.IntVar(frame2, self.lMsSink)
        radio4 = tk.Radiobutton(frame2, text="Threshold Model", variable=var3, value=0, width=35, anchor='w')
        radio5 = tk.Radiobutton(frame2, text="S-Shaped Model", variable=var3, value=1, width=35, anchor='w')
        
        radio1.pack()
        radio2.pack()
        radio3.pack()
        radio4.pack()
        radio5.pack()
        
        tk.Label(frame3, text="Active Solute Uptake Model", width=50, anchor='w').pack(pady=10)
        var4 = tk.IntVar(canvas2, self.lActRSU)
        Check1 = tk.Checkbutton(canvas2, text="Active Solute Uptake", variable=var4, width=30, anchor='w')
        Check1.place(x=5, y=0)
        tk.Label(canvas2, text="Potental Solute Uptake Rate", width=30, anchor='w').place(x=5, y=30)
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, 0)
        e2.place(x=250, y=30)
        tk.Label(canvas2, text="Michaelis-Menten Constant", width=30, anchor='w').place(x=5, y=60)
        e3 = tk.Entry(canvas2, width=10, justify='right')
        e3.insert(0, 0)
        e3.place(x=250, y=60)
        tk.Label(canvas2, text="Minimum Concentration for Uptake", width=30, anchor='w').place(x=5, y=90)
        e4 = tk.Entry(canvas2, width=10, justify='right')
        e4.insert(0, 0)
        e4.place(x=250, y=90)
        tk.Label(canvas2, text="Critical Solute Index", width=30, anchor='w').place(x=5, y=120)
        e5 = tk.Entry(canvas2, width=10, justify='right')
        e5.insert(0, 1)
        e5.place(x=250, y=120)
        
        var5 = tk.IntVar(canvas2, 1)
        Check1 = tk.Checkbutton(canvas2, text="Reduced Potential Solute Uptake\ndue to Reduced Water Uptake", variable=var5, width=30, anchor='w')
        Check1.place(x=5, y=150)

        #if self.lChem:
        for child in canvas2.winfo_children():
            child.configure(state='disable')
        
        canvas2.pack(padx=10, pady=0)

        def _GetValues():
            values = {
                'rootModel': bool(var1.get()),
                'OmegaC': e1.get(),
                'lActRSU':  bool(var4.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if not self.rootModel:
                self.Root_Water_Uptake_Parameters1()
            else:
                self.Root_Water_Uptake_Parameters2()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lTemp:
                self.Heat_Transport_Parameters()
            if self.lWTDep:
                self.Water_Dependent_Solute_Reaction_Parameters()
            elif self.lTDep:
                self.Temperature_Dependent_Solute_Transport_and_Reaction_Parameters()
            elif self.lAddFum:
                self.Add_Fumigant()
            elif self.lChem:
                self.Reqaction_Parameters_for_solute()
            elif self.lWat:
                self.Water_Flow_Parameters()

        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0)
        frame2.grid(row=1, column=0)
        frame3.grid(row=2, column=0)
        frame4.grid(row=0, column=1, rowspan=2)

        root.mainloop()
        
    def Root_Water_Uptake_Parameters1(self):
        root = tk.Tk()
        root.title("Root Water Uptake Parameters")
        root.geometry("400x300")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')

        tk.Label(frame1, text="Feddes\' parameters", width=30, anchor='w').pack(pady=10)
        tk.Label(canvas1, text="PO", width=10, anchor='w').pack()
        tk.Label(canvas1, text="P2H", width=10, anchor='w').pack()
        tk.Label(canvas1, text="P2L", width=10, anchor='w').pack()
        tk.Label(canvas1, text="P3", width=10, anchor='w').pack()
        tk.Label(canvas1, text="r2H", width=10, anchor='w').pack()
        tk.Label(canvas1, text="r2L", width=10, anchor='w').pack()
        tk.Label(canvas1, text="POpt", width=10, anchor='w').pack()

        e1 = tk.Entry(canvas2, width=10, justify='right')
        e1.insert(0, self.uptake_param[0])
        e1.pack()
        e2 = tk.Entry(canvas2, width=10, justify='right')
        e2.insert(0, self.uptake_param[1])
        e2.pack()
        e3 = tk.Entry(canvas2, width=10, justify='right')
        e3.insert(0, self.uptake_param[2])
        e3.pack()
        e4 = tk.Entry(canvas2, width=10, justify='right')
        e4.insert(0, self.uptake_param[3])
        e4.pack()
        e5 = tk.Entry(canvas2, width=10, justify='right')
        e5.insert(0, self.uptake_param[4])
        e5.pack()
        e6 = tk.Entry(canvas2, width=10, justify='right')
        e6.insert(0, self.uptake_param[5])
        e6.pack()
        e7 = tk.Entry(canvas2, width=10, justify='right')
        e7.insert(0, self.uptake_param[6])
        e7.pack()
        
        canvas1.pack(side='left', padx=10)
        canvas2.pack(padx=10)
        
        def ChangeCrop(event):
            Crop = var1.get()
            for child in canvas2.winfo_children():
                child.delete(0, 'end')
            for i, child in enumerate(canvas2.winfo_children()):
                child.insert(0, self.Uptake_catalog[Crop][i])
            
            
        tk.Label(frame2, text="Database", width=30, anchor='w').pack(pady=10)
        var1 = tk.StringVar(frame2, self.croptype)
        Combo1 = ttk.Combobox(frame2, width=10, textvariable=var1)
        Combo1['values'] = list(self.Uptake_catalog.keys())
        Combo1.current()
        Combo1.bind("<<ComboboxSelected>>", ChangeCrop)
        Combo1.pack()
        
        def _GetValues():
            values = {
                'uptake_param': [float(child.get()) for child in canvas2.winfo_children()],
                'croptype': var1.get(),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lMsSink:
                self.Root_Water_Uptake_Parameters3()
            elif self.lRootGr:
                self.Root_Growth_Parameters()
                
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Root_Water_and_Solute_Uptake_Model()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame3, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame3, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame3, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame3, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=10)
        frame2.grid(row=1, column=0, padx=10)
        frame3.grid(row=0, column=1, rowspan=2, padx=10)
        
        root.mainloop()

    def Root_Water_Uptake_Parameters2(self):
        root = tk.Tk()
        root.title("Root Water Uptake Parameters")
        root.geometry("300x200")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')

        tk.Label(frame1, text="S-Shape parameters", width=30, anchor='w').grid(row=0, column=0, columnspan=2)
        tk.Label(frame1, text="P5O", width=5, anchor='w').grid(row=1, column=0)
        tk.Label(frame1, text="P3", width=5, anchor='w').grid(row=2, column=0)
        tk.Label(frame1, text="PW", width=5, anchor='w').grid(row=3, column=0)

        e1 = tk.Entry(frame1, width=10, justify='right')
        e1.insert(0, self.S_shape_Param[0])
        e1.grid(row=1, column=1)
        e2 = tk.Entry(frame1, width=10, justify='right')
        e2.insert(0, self.S_shape_Param[1])
        e2.grid(row=2, column=1)
        e3 = tk.Entry(frame1, width=10, justify='right')
        e3.insert(0, self.S_shape_Param[2])
        e3.grid(row=3, column=1)

        
        def _GetValues():
            values = {
                'S_shape_Param':  [e1.get(), e2.get(), e3.get(),],
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lMsSink:
                self.Root_Water_Uptake_Parameters3()
            elif self.lRootGr:
                self.Root_Growth_Parameters()
                
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self.Root_Water_and_Solute_Uptake_Model()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame2, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame2, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame2, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame2, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=10)
        frame2.grid(row=0, column=1, rowspan=2, padx=10)
        
        root.mainloop()

    def Root_Water_Uptake_Parameters3(self):
        root = tk.Tk()
        root.title("Root Water Uptake Parameters")
        root.geometry("400x500")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame2)
        frame3 =tk.Frame(root, bd=2, relief='groove')
        frame4 =tk.Frame(root, bd=2, relief='groove')
        
        if self.lMsSink:
            tk.Label(frame1, text="S-Shape Model", width=30, anchor='w').grid(row=0, column=0, columnspan=2)
        else:
            tk.Label(frame1, text="Threshold Model", width=30, anchor='w').grid(row=0, column=0, columnspan=2)
            
        tk.Label(frame1, text="Threshold:", width=10, anchor='w').grid(row=1, column=0)
        tk.Label(frame1, text="Slope:", width=10, anchor='w').grid(row=2, column=0)

        e1 = tk.Entry(frame1, width=10, justify='right')
        e1.insert(0, self.Threshold_Model_param[0])
        e1.grid(row=1, column=1)
        e2 = tk.Entry(frame1, width=10, justify='right')
        e2.insert(0, self.Threshold_Model_param[1])
        e2.grid(row=2, column=1)
        
        
        tk.Label(frame2, text="Conversion to pressure/osmotic head", width=30, anchor='w').pack()
        header = ('', 'Osm.Coeff.')
        trv1 = Tableview(canvas1, columns=tuple(range(len(header))), show='headings')
        for i, key in enumerate(header):
            trv1.column(i, anchor='center', width=100)
            trv1.heading(i, text=key)
    
        for i, row in enumerate(self.aOsm):
            trv1.insert('', tk.END, values=[i+1, row])
        
        scrollbar1 = ttk.Scrollbar(canvas1, orient='vertical', command=trv1.yview)
        scrollbar1.pack(side ='right', fill ='y')
        trv1.configure(yscrollcommand = scrollbar1.set)
        trv1.pack(padx=5, pady=5)

        canvas1.pack(padx=10, pady=10)
        
        def ChangeCrop(event):
            croptype_threshold = var1.get()
            Threshold_Model_param = self.Threshold_Model_catalog[croptype_threshold]
            e1.delete(0, tk.END)
            e2.delete(0, tk.END)
            e1.insert(0, Threshold_Model_param[0])
            e2.insert(0, Threshold_Model_param[1])
            
        tk.Label(frame3, text="Database", width=30, anchor='w').pack(pady=10)
        var1 = tk.StringVar(frame3, self.croptype_threshold)
        Combo1 = ttk.Combobox(frame3, width=10, textvariable=var1)
        Combo1['values'] = list(self.Threshold_Model_catalog.keys())
        Combo1.current()
        Combo1.bind("<<ComboboxSelected>>", ChangeCrop)
        Combo1.pack()
        
        if self.SoluteAdditive:
            for child in frame1.winfo_children():
                child.config(state='disabled')
            for child in frame3.winfo_children():
                child.config(state='disabled')
        else:
            for child in frame1.winfo_children():
                child.config(state='normal')
            for child in frame3.winfo_children():
                child.config(state='normal')
        
        def _GetValues():
            values = {
                'Threshold_Model_param': [float(e1.get()), float(e2.get())],
                'aOsm': trv1.get(0, 'end'),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lRootGr:
                self.Root_Growth_Parameters()
                
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if not self.rootModel:
                self.Root_Water_Uptake_Parameters1()
            else:
                self.Root_Water_Uptake_Parameters2()            
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame4, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame4, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame4, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame4, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=10, pady=10)
        frame2.grid(row=1, column=0, padx=10, pady=10)
        frame3.grid(row=2, column=0, padx=10, pady=10)
        frame4.grid(row=0, column=1, rowspan=3, padx=10, pady=10)
        
        root.mainloop()


    def Root_Growth_Parameters(self):
        root = tk.Tk()
        root.title("Root Water Uptake Parameters")
        root.geometry("700x500")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')
        frame3 =tk.Frame(root, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame3)
        canvas2 = tk.Canvas(frame3)
        frame4 =tk.Frame(root, bd=2, relief='groove')
        canvas3 = tk.Canvas(frame4)
        canvas4 = tk.Canvas(frame4)
        frame5 =tk.Frame(root, bd=2, relief='groove')        
        rCenter = self.rCenter

        def __Check_1():
            if var3.get():
                for child in canvas2.winfo_children():
                    child.destroy()
                header = ['', 'Plant Position']
                trv2 = Tableview(canvas2, columns=tuple(range(len(header))), show='headings')
                for i, key in enumerate(header):
                    trv2.column(i, anchor='center', width=100)
                    trv2.heading(i, text=key)
                for i, row in enumerate(rCenter):
                    trv2.insert('', tk.END, values=[i+1, row])
                trv2.pack(padx=5, pady=5)    
                canvas2.pack()
                
                for child in canvas1.winfo_children():
                    child.config(state='normal')
                if var1.get() == 0 and var2.get() != 0:
                    e1.config(state='disabled')
                    e2.config(state='disabled')
            else:
                for child in canvas1.winfo_children():
                    child.config(state='disabled')
                for child in canvas2.winfo_children():
                    child.destroy()
                if var1.get() == 1 and var2.get() == 0:
                    e1.config(state='normal')

        def Root_Growth_Data():
            if var2.get() == 0:
                canvas3.pack_forget()   
                canvas4.pack_forget()
            elif var2.get() == 1:
                canvas4.pack_forget()
                canvas3.pack()
            elif var2.get() == 2:
                canvas3.pack_forget()   
                canvas4.pack()
            __Check_1()

        tk.Label(frame1, text="Shape of the Root Zone", width=30, anchor='w').pack()
        var1  = tk.IntVar(frame1, self.iRootZoneShape)
        radio1 = tk.Radiobutton(frame1, text="Vrugt et al. (2000)", variable=var1, value=0, width=30, anchor='w', command=Root_Growth_Data)
        radio2 = tk.Radiobutton(frame1, text="van Genuchten & Hoffman", variable=var1, value=1, width=30, anchor='w', command=Root_Growth_Data)
        radio3 = tk.Radiobutton(frame1, text="Constant", variable=var1, value=2, width=30, anchor='w', command=Root_Growth_Data)
        radio1.pack()
        radio2.pack()
        radio3.pack()
        
        tk.Label(frame2, text="Root Depth Specified", width=40, anchor='w').pack()
        var2  = tk.IntVar(frame2, self.iRootDepthEntry)
        radio1 = tk.Radiobutton(frame2, text="No Growth", variable=var2, value=0, width=40, anchor='w', command=Root_Growth_Data)
        radio2 = tk.Radiobutton(frame2, text="Using a Table", variable=var2, value=1, width=40, anchor='w', command=Root_Growth_Data)
        radio3 = tk.Radiobutton(frame2, text="Using a Growth Function", variable=var2, value=2, width=40, anchor='w', command=Root_Growth_Data)
        radio1.pack()
        radio2.pack()
        radio3.pack()
        
        def _Update():
            NPlants = int(e3.get())
            trv2 = canvas2.winfo_children()[0]
            rCenter = [0 for _ in range(NPlants)]
            for item in trv2.get_children():
                trv2.delete(item)
            for i in range(NPlants):
                trv2.insert('', tk.END, values=[i+1, rCenter[i]])
        
        tk.Label(frame3, text="Horizontal Distribution of Plants", width=30, anchor='w').pack(pady=10)
        var3 = tk.IntVar(frame3, self.Horizontal)
        Check1 = tk.Checkbutton(frame3, text="Horizontal Distribution", variable=var3, width=30, anchor='w', command=__Check_1)
        Check1.pack()
        tk.Label(canvas1, text="Root Depth:", width=15, anchor='w').grid(row=0, column=0)
        e1 = tk.Entry(canvas1, width=10, justify='right')
        e1.insert(0, self.rDepth)
        e1.grid(row=0, column=1)
        tk.Label(canvas1, text="Half Width", width=15, anchor='w').grid(row=1, column=0)
        e2 = tk.Entry(canvas1, width=10, justify='right')
        e2.insert(0, self.RootHalfWidth)
        e2.grid(row=1, column=1)
        tk.Label(canvas1, text="Number of Plants", width=15, anchor='w').grid(row=2, column=0)
        e3 = tk.Entry(canvas1, width=10, justify='right')
        e3.insert(0, self.NPlants)
        e3.grid(row=2, column=1)
        button1 = tk.Button(canvas1, text='_Update', command=_Update)
        button1.grid(row=2, column=2)

        canvas1.pack()
        
        tk.Label(frame4, text="Root Growth Data", width=50, anchor='w').pack(pady=10)

        def _Update_RootDepth():
            nGrowth = float(e14.get())
            time = np.arange(0, nGrowth)
            RootDepth = [[t, 0] for t in time]
            for item in trv1.get_children():
                trv1.delete(item)
            for i, row in enumerate(RootDepth):
                trv1.insert('', tk.END, values=[i+1]+row)
    
        header = ['', 'Time', 'Depth']
        tk.Label(canvas3, text='Number of Rows:', width=15, anchor='w').grid(row=0, column=0)
        e14 = tk.Entry(canvas3, width=10, justify='right')
        e14.insert(0, self.nGrowth)
        e14.grid(row=0, column=1)
        tk.Button(canvas3, text='_Update', command=_Update_RootDepth).grid(row=0, column=2)
        trv1 = Tableview(canvas3, columns=tuple(range(len(header))), show='headings')
        for i, key in enumerate(header):
            trv1.column(i, anchor='center', width=100)
            trv1.heading(i, text=key)
    
        for i, row in enumerate(self.RootDepth):
            trv1.insert('', tk.END, values=[i+1]+row)
        trv1.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

        def _Radio_1():
            if var4.get():
                e9.config(state='normal')
                e10.config(state='normal')
            else:
                e9.config(state='disabled')
                e10.config(state='disabled')

        tk.Label(canvas4, text="Root Growth Factor:", width=40, anchor='w').grid(row=0, column=0, columnspan=2)
        var4  = tk.IntVar(canvas4, self.iRFak)
        radio1 = tk.Radiobutton(canvas4, text="From Given Data", variable=var4, value=0, width=40, anchor='w', command=_Radio_1)
        radio2 = tk.Radiobutton(canvas4, text="50% after 50% Growing Season", variable=var4, value=1, width=40, anchor='w', command=_Radio_1)
        radio1.grid(row=1, column=0, columnspan=2)
        radio2.grid(row=2, column=0, columnspan=2)

        #ttk.Separator(canvas4, orient='horizontal').pack(fill='x', padx=10, pady=5)
        
        tk.Label(canvas4, text="Inital Root Growth Time: [{}]".format(self.TUnit), width=30, anchor='w').grid(row=3, column=0)
        tk.Label(canvas4, text="Harvest Time: [{}]".format(self.TUnit), width=30, anchor='w').grid(row=4, column=0)
        tk.Label(canvas4, text="Inital Root Depth: [{}]".format(self.LUnit), width=30, anchor='w').grid(row=5, column=0)
        tk.Label(canvas4, text="Maximum Root Depth: [{}]".format(self.LUnit), width=30, anchor='w').grid(row=6, column=0)
        tk.Label(canvas4, text="'Inital Rooting Radius: [{}]".format(self.LUnit), width=30, anchor='w').grid(row=7, column=0)
        tk.Label(canvas4, text="Maximum Rooting Radius: [{}]".format(self.LUnit), width=30, anchor='w').grid(row=8, column=0)
        tk.Label(canvas4, text="Time - Root Data: [{}]".format(self.TUnit), width=30, anchor='w').grid(row=9, column=0)
        tk.Label(canvas4, text="Depth - Root Data: [{}]".format(self.LUnit), width=30, anchor='w').grid(row=10, column=0)
        tk.Label(canvas4, text="Time-Period: [{}]".format(self.TUnit), width=30, anchor='w').grid(row=11, column=0)

        e5 = tk.Entry(canvas4, width=10, justify='right')
        e5.insert(0, self.tRMin)
        e5.grid(row=3, column=1)
        e6 = tk.Entry(canvas4, width=10, justify='right')
        e6.insert(0, self.tRMax)
        e6.grid(row=4, column=1)
        e7 = tk.Entry(canvas4, width=10, justify='right')
        e7.insert(0, self.ZRMin)
        e7.grid(row=5, column=1)
        e8 = tk.Entry(canvas4, width=10, justify='right')
        e8.insert(0, self.ZRMax)
        e8.grid(row=6, column=1)
        e9 = tk.Entry(canvas4, width=10, justify='right')
        e9.insert(0, 0.01)
        e9.grid(row=7, column=1)
        e10 = tk.Entry(canvas4, width=10, justify='right')
        e10.insert(0, 45)
        e10.grid(row=8, column=1)
        e11 = tk.Entry(canvas4, width=10, justify='right')
        e11.insert(0, self.tRMed)
        e11.grid(row=9, column=1)
        e12 = tk.Entry(canvas4, width=10, justify='right')
        e12.insert(0, self.ZRMed)
        e12.grid(row=10, column=1)
        e13 = tk.Entry(canvas4, width=10, justify='right')
        e13.insert(0, self.tPeriodRoot)
        e13.grid(row=11, column=1)

        Root_Growth_Data()
        __Check_1()

        def _GetValues():
            values = {
                'iRootZoneShape':   bool(var1.get()),
                'iRootDepthEntry':  bool(var2.get()),
                'Horizontal':       bool(var3.get()),
                'iRFak':            bool(var4.get()),
                'rDepth':        float(e1.get()),
                'RootHalfWidth':    float(e2.get()),
                'NPlants':          int(e3.get()),
                'tRMin':            float(e5.get()),
                'tRMax':            float(e6.get()),
                'ZRMin':            float(e7.get()),
                'ZRMax':            float(e8.get()),
                'tRMed':            float(e11.get()),
                'ZRMed':            float(e12.get()),
                'tPeriodRoot':      float(e13.get()),
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lTDep:
                self.Time_Variable_Boundary_Conditions()
            else:
                self._Update_Tree()
                
        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            if self.lMsSink:
                self.Root_Water_Uptake_Parameters3()            
            elif not self.rootModel:
                self.Root_Water_Uptake_Parameters1()
            else:
                self.Root_Water_Uptake_Parameters2()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()
            self._Update_Tree()

        button__OK = tk.Button(frame5, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame5, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame5, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame5, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame1.grid(row=0, column=0, padx=5)
        frame2.grid(row=0, column=1, padx=5)
        frame3.grid(row=1, column=0, padx=5)
        frame4.grid(row=1, column=1, padx=5)
        frame5.grid(row=0, column=2, rowspan=2, padx=10)
        
        root.mainloop()


    def Time_Variable_Boundary_Conditions(self):
        root = tk.Tk()
        root.title("Time Variable Boundary Conditions")
        root.geometry("1000x500")
        frame1 =tk.Frame(root, bd=2, relief='groove')
        frame2 =tk.Frame(root, bd=2, relief='groove')

        header = ['Time', 'Precip.', 'Evap', 'Transp.', 'hCritA', 'Var.Fl1', 'Var.H-1', 'Var.Fl2', 'Var.H-2', 'Var.Fl3', 'Var.H-3', 'Var.Fl4', 'Var.H-4']

        tk.Label(frame1, text="Parameters", width=70, anchor='w').pack()
        trv1 = Tableview(frame1, columns=tuple(range(len(header))), show='headings')
        for i, key in enumerate(header):
            trv1.column(i, anchor='center', width=60)            
            trv1.heading(i, text=key)
        
        for row in self.atmosph_data:
            trv1.insert('', tk.END, values=row)
        trv1.pack()
        
        frame1.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

        tk.Button(root, text='linear relation of time between the initial and final time', width=50).grid(row=1, column=0, padx=10, pady=10)
        tk.Label(root, text='Surface area associated with transpiration: ').grid(row=2, column=0, padx=10, pady=10)
        e1 = tk.Entry(root, justify='right', width=10)
        e1.insert(0, self.hCritS)
        e1.grid(row=2, column=1, padx=10, pady=10)
        
        def _GetValues():
            values = {
                'hCritS': float(e1.get()),
                'atmosph_data': [trv1.item(line)['values'] for line in trv1.get_children()]
            }
            return values

        def _Next():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()

        def _Previous():
            self.SaveNewValues(_GetValues())
            root.destroy()
            
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
            self._Update_Tree()
        
        def _Cancel():
            root.destroy()        
            self._Update_Tree()

        button__OK = tk.Button(frame2, text="OK", command=_OK, width=7)
        button_C = tk.Button(frame2, text="Cancel", command=_Cancel, width=7)
        button__Next = tk.Button(frame2, text="Next", command=_Next, width=7)
        button_Pre = tk.Button(frame2, text="Previous", command=_Previous, width=7)
        button__OK.pack()
        button_C.pack()
        button__Next.pack()
        button_Pre.pack()

        frame2.grid(row=0, column=3, rowspan=3, padx=10, pady=10)

        root.mainloop()         


    def Boundary_Condition_Options(self):
        root = tk.Tk()
        root.title("Boundary Condition Options")
        root.geometry("600x450")
        tabRoot = ttk.Notebook(root)
        frame1 =tk.Frame(root)
          
        tab1 = ttk.Frame(tabRoot)
        tab2 = ttk.Frame(tabRoot)
        tab3 = ttk.Frame(tabRoot)
        tab4 = ttk.Frame(tabRoot)
          
        tabRoot.add(tab1, text ='Time-Variable Head/Flux 1')
        tabRoot.add(tab2, text ='Special BC')
        tabRoot.add(tab3, text ='Triggered Irrigation')
        tabRoot.add(tab4, text ='Reservoir Boundary Condition')
        tabRoot.pack(expand = 1, fill ="both", padx=5, pady=5)
       
        self._Boundary_Condition_Options_tab1(tab1)
        self._Boundary_Condition_Options_tab2(tab2)
        self._Boundary_Condition_Options_tab3(tab3)
        self._Boundary_Condition_Options_tab4(tab4)
        
        def _GetValues():
            Variable = {
            #tab1
                'Interp':   bool(self.var1_1.get()),
                'H_Flux':   bool(self.var1_2.get()),
                'H_Flx1':   bool(self.var1_3.get()),
                'Atm_H':    bool(self.var1_4.get()),
                'Seep_H':   bool(self.var1_5.get()),
                'Atm_WL':   bool(self.var1_6.get()),
                'Atm_SF':   bool(self.var1_7.get()),
                'Snow':     bool(self.var1_8.get()),
            #tab2
                'Gradient': bool(self.var2_1.get()),
                'SubDrip':  bool(self.var2_3.get()),
                'SurfDrip': bool(self.var2_4.get()),
                'iDripNode':    self.var2_5.get(),
                'SeepFace': bool(self.var2_6.get()),
                'xGrad':        self.e2_1.get(),
                'QDrip':        self.e2_2.get(),
                'ExpDrip':      self.e2_3.get(),
                'iDripCenter':  self.e2_4.get(),
                'hSeep':        self.e2_5.get(),
            #tab3
                'TriggIrrig':   bool(self.var3_1.get()),
                'kIrrig':       self.var3_2.get(),
                'iIrrig':       self.e3_1.get(),
                'hIrrig':       self.e3_2.get(),
                'rIrrig':       self.e3_3.get(),
                'tIrrig':       self.e3_4.get(),
                'lIrrig':       self.e3_5.get(),
            #tab4
                'WellBC':   bool(self.var4_1.get()),
                'iWell':    self.var2.get(),
                'zWBot':    self.e4_1.get(),
                'Radius':   self.e4_2.get(),
                'WellPump': self.e4_3.get(),
                'zWInit':   self.e4_4.get(),
                'cWell':    self.e4_5.get(),
            }
            return Variable
        
        def _OK():
            self.SaveNewValues(_GetValues())
            root.destroy()
        
        def _Cancel():
            root.destroy()
            
        button__OK = tk.Button(frame1, text="OK", command=_OK, width=7)
        button__Cancel = tk.Button(frame1, text="Cancel", command=_Cancel, width=7)
        button__OK.pack(side='left', padx=5)
        button__Cancel.pack(side='left', anchor='e', padx=5)
        frame1.pack(expand=1, anchor='e', padx=5, pady=5)
        root.mainloop()
        
        
    def _Boundary_Condition_Options_tab1(self, tab):
        frame1 = tk.Frame(tab, bd=2, relief='groove')
        frame2 = tk.Frame(tab, bd=2, relief='groove')
        
        tk.Label(frame1, text='Basic Options', anchor='w', width=70).pack()
        self.var1_1 = tk.IntVar(frame1, self.Interp)
        Check1 = tk.Checkbutton(frame1, text='Interpolate Time-Variable Pressure Head and flux(1) boundary conditions in time', variable=self.var1_1, anchor='w', width=70)
        Check1.pack()
        
        ttk.Separator(frame1, orient='horizontal').pack(fill='x', padx=10)
        
        def __Check_Baskc_Options():
            Vars = [self.var1_2.get(), self.var1_3.get(), self.var1_4.get(), self.var1_5.get(), self.var1_6.get(), self.var1_7.get()]
            Checks = [Check2, Check3, Check4, Check5, Check6, Check7]
            if sum(Vars):
                for i, var in enumerate(Vars):
                    Checks[i].config(state='disabled')
                    if var:
                        Checks[i].config(state='normal')
            else:
                for check in Checks:
                    check.config(state='normal')

        tk.Label(frame1, text='Switch the boundary condition from Time-Variable Pressure Head to:', anchor='w', width=70).pack()
        self.var1_2 = tk.IntVar(frame1, self.H_Flux)
        Check2 = tk.Checkbutton(frame1, text='NO Flux BC when Var-H.1>999999', variable=self.var1_2, anchor='w', width=60, command=__Check_Baskc_Options)
        Check2.pack()
        self.var1_3 = tk.IntVar(frame1, self.H_Flx1)
        Check3 = tk.Checkbutton(frame1, text='NO Flux BC when the specified nodal pressure head is negative', variable=self.var1_3, anchor='w', width=60, command=__Check_Baskc_Options)
        Check3.pack()
        self.var1_4 = tk.IntVar(frame1, self.Atm_H)
        Check4 = tk.Checkbutton(frame1, text='Atmospheric BC when the specified nodal pressure head is negative', variable=self.var1_4, anchor='w', width=60, command=__Check_Baskc_Options)
        Check4.pack()
        self.var1_5 = tk.IntVar(frame1, self.Seep_H)
        Check5 = tk.Checkbutton(frame1, text='Seepage Face BC when the specified nodal pressure head is negative', variable=self.var1_5, anchor='w', width=60, command=__Check_Baskc_Options)
        Check5.pack()

        ttk.Separator(frame1, orient='horizontal').pack(fill='x', padx=10)

        self.var1_6 = tk.IntVar(frame1, self.Atm_WL)
        Check6 = tk.Checkbutton(frame1, text='Treat the Time-Variable Flux boundary condition as the Atmospheric BC\n(i.e., with limited pressure heads)', variable=self.var1_6, anchor='w', width=70, command=__Check_Baskc_Options)
        Check6.pack()
        self.var1_7 = tk.IntVar(frame1, self.Atm_SF)
        Check7 = tk.Checkbutton(frame1, text='Apply Atmospheric boundary condition to nonactive Seepage Face', variable=self.var1_7, anchor='w', width=70, command=__Check_Baskc_Options)
        Check7.pack()
        
        tk.Label(frame2, text='Other Boundary Conditions Options', anchor='w', width=70).pack()
        self.var1_8 = tk.IntVar(frame2, self.Snow)
        Check8 = tk.Checkbutton(frame2, text='Consider snow accumulation at the soil surface when temperatures are negative', variable=self.var1_8, anchor='w', width=70)
        Check8.pack()
        self.var1_9 = tk.IntVar(frame2, 0)
        Check9 = tk.Checkbutton(frame2, text='Consider only horizontal projections of boundary elements for Time-Variable Flux BCs', variable=self.var1_9, anchor='w', width=70)
        Check9.pack()
        
        Check8.config(state='disabled')
        frame1.pack(pady=5)
        frame2.pack(pady=5)
        
        
    def _Boundary_Condition_Options_tab2(self, tab):
        frame1 = tk.Frame(tab, bd=2, relief='groove')
        frame2 = tk.Frame(tab, bd=2, relief='groove')
        frame3 = tk.Frame(tab, bd=2, relief='groove')
        frame4 = tk.Frame(tab, bd=2, relief='groove')
        
        def __Check_1():
            if self.var2_1.get():
                self.e2_1.config(state='normal')
            else: self.e2_1.config(state='disabled')

        tk.Label(frame1, text='Gradient Boundary Conditions (instead of Free Drainage BC)', anchor='w', width=70).grid(row=0, column=0, columnspan=3)
        self.var2_1 = tk.IntVar(frame1, self.Gradient)
        Check1 = tk.Checkbutton(frame1, text='Gradient in the x-direction (positive against the x-axis)', variable=self.var2_1, anchor='w', width=50, command=__Check_1)
        Check1.grid(row=1, column=0)
        self.var2_2 = tk.IntVar(frame1, 0)
        Check2 = tk.Checkbutton(frame1, text='Gradient in the y-direction (positive against the y-axis)', variable=self.var2_2, anchor='w', width=50)
        Check2.config(state='disabled')
        Check2.grid(row=2, column=0)
        tk.Label(frame1, text='Gradient:', anchor='w', width=10).grid(row=1, column=1)
        self.e2_1 = tk.Entry(frame1, justify='left', width=10)
        self.e2_1.insert(0, self.xGrad)
        self.e2_1.grid(row=1, column=2)
        __Check_1()
        
        def __Check_3():
            if self.var2_3.get():
                self.e2_2.config(state='normal')
                self.e2_3.config(state='normal')
            else:
                self.e2_2.config(state='disabled')
                self.e2_3.config(state='disabled')

        tk.Label(frame2, text='Subsurface Drip Characteristic Function (for Time-Variable Flux 1 BC)', anchor='w', width=70).grid(row=0, column=0, columnspan=3)
        self.var2_3 = tk.IntVar(frame2, self.SubDrip)
        Check3 = tk.Checkbutton(frame2, text='Dripper Characterstic Function', variable=self.var2_3, anchor='w', width=50, command=__Check_3)
        Check3.grid(row=1, column=0)
        tk.Label(frame2, text='Opt. Flux:', anchor='w', width=10).grid(row=1, column=1)
        self.e2_2 = tk.Entry(frame2, justify='left', width=10)
        self.e2_2.insert(0, self.QDrip)
        self.e2_2.grid(row=1, column=2)
        tk.Label(frame2, text='Exponent:', anchor='w', width=10).grid(row=2, column=1)
        self.e2_3 = tk.Entry(frame2, justify='left', width=10)
        self.e2_3.insert(0, self.ExpDrip)
        self.e2_3.grid(row=2, column=2)
        __Check_3()

        def __Check_4():
            if self.var2_4.get():
                radio1.config(state='normal')
                radio2.config(state='normal')
                radio3.config(state='normal')
                self.e2_4.config(state='normal')
            else:
                radio1.config(state='disabled')
                radio2.config(state='disabled')
                radio3.config(state='disabled')
                self.e2_4.config(state='disabled')

        tk.Label(frame3, text='Surface Drip Characteristic Function (for Time-Variable Flux 1 BC)', anchor='w', width=70).grid(row=0, column=0, columnspan=4)
        self.var2_4 = tk.IntVar(frame3, self.SurfDrip)
        Check4 = tk.Checkbutton(frame3, text='Surface Drip with Dynamic Wetting', variable=self.var2_4, anchor='w', width=30, command=__Check_4)
        Check4.grid(row=1, column=0)
        self.var2_5 = tk.IntVar(frame3, self.iDripNode)
        radio1 = tk.Radiobutton(frame3, text='From left to right', variable=self.var2_5, value=-1, anchor='w', width=15)
        radio2 = tk.Radiobutton(frame3, text='From the center', variable=self.var2_5, value=0, anchor='w', width=15)
        radio3 = tk.Radiobutton(frame3, text='From right to left', variable=self.var2_5, value=1, anchor='w', width=15)
        radio1.grid(row=1, column=1)
        radio2.grid(row=2, column=1)
        radio3.grid(row=3, column=1)
        tk.Label(frame3, text='Center Node:', anchor='w', width=11).grid(row=1, column=2, rowspan=3)
        self.e2_4 = tk.Entry(frame3, justify='left', width=10)
        self.e2_4.insert(0, self.iDripCenter)
        self.e2_4.grid(row=1, column=3, rowspan=3)
        __Check_4()
        
        def __Check_5():
            if self.var2_6.get():
                self.e2_5.config(state='normal')
            else:
                self.e2_5.config(state='disabled')

        tk.Label(frame4, text='Seepage Face', anchor='w', width=70).grid(row=0, column=0, columnspan=3)
        self.var2_6 = tk.IntVar(frame4, self.SeepFace)
        Check5 = tk.Checkbutton(frame4, text='Seepage Face with Specified Pressure Head', variable=self.var2_6, anchor='w', width=45, command=__Check_5)
        Check5.grid(row=1, column=0)
        tk.Label(frame4, text='Pressure Head:', anchor='w', width=15).grid(row=1, column=1)
        self.e2_5 = tk.Entry(frame4, justify='left', width=10)
        self.e2_5.insert(0, self.hSeep)
        self.e2_5.grid(row=1, column=2)
        
        frame1.pack(pady=5)
        frame2.pack(pady=5)
        frame3.pack(pady=5)
        frame4.pack(pady=5)



    def _Boundary_Condition_Options_tab3(self, tab):
        frame1 = tk.Frame(tab, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        
        tk.Label(frame1, text='Triggered Irrigation (by Specified Pressure Head in a Observation Node)', anchor='w', width=70).pack()
        
        def __Check_1():
            if self.var3_1.get():
                for child in canvas1.winfo_children():
                    child.config(state='normal')
            else:
                for child in canvas1.winfo_children():
                    child.config(state='disabled')

        self.var3_1 = tk.IntVar(frame1, self.TriggIrrig)
        Check1 = tk.Checkbutton(frame1, text='Triggered Irrigation', variable=self.var3_1, anchor='w', width=70, command=__Check_1)
        Check1.pack()
        
        tk.Label(canvas1, text='Observation Node Triggering Irrigation:', anchor='w', width=40).grid(row=0, column=0)
        tk.Label(canvas1, text='Pressure Head Triggering Irrigation:', anchor='w', width=40).grid(row=1, column=0)
        self.e3_1 = tk.Entry(canvas1, justify='left', width=10)
        self.e3_1.insert(0, self.iIrrig)
        self.e3_1.grid(row=0, column=1)
        self.e3_2 = tk.Entry(canvas1, justify='left', width=10)
        self.e3_2.insert(0, self.hIrrig)
        self.e3_2.grid(row=1, column=1)
        
        tk.Label(canvas1, text='Boundary where irrigation will be applied:', anchor='w', width=40).grid(row=2, column=0)
        self.var3_2 = tk.IntVar(canvas1, self.kIrrig)
        radio1 = tk.Radiobutton(canvas1, text='Variable Flux Boundary', variable=self.var3_2, value=-3, anchor='w', width=40)
        radio2 = tk.Radiobutton(canvas1, text='Variable Head Boundary', variable=self.var3_2, value=3, anchor='w', width=40)
        radio3 = tk.Radiobutton(canvas1, text='Atmospheric Boundary', variable=self.var3_2, value=-4, anchor='w', width=40)
        radio1.grid(row=3, column=0)
        radio2.grid(row=4, column=0)
        radio3.grid(row=5, column=0)

        tk.Label(canvas1, text='Irrigation Rate:', anchor='w', width=20).grid(row=3, column=1)
        tk.Label(canvas1, text='Irrigation Duration:', anchor='w', width=20).grid(row=4, column=1)
        tk.Label(canvas1, text='Lag Time:', anchor='w', width=20).grid(row=5, column=1)
        self.e3_3 = tk.Entry(canvas1, justify='left', width=10)
        self.e3_3.insert(0, self.rIrrig)
        self.e3_3.grid(row=3, column=2)
        self.e3_4 = tk.Entry(canvas1, justify='left', width=10)
        self.e3_4.insert(0, self.tIrrig)
        self.e3_4.grid(row=4, column=2)
        self.e3_5 = tk.Entry(canvas1, justify='left', width=10)
        self.e3_5.insert(0, self.lIrrig)
        self.e3_5.grid(row=5, column=2)
        
        canvas1.pack(padx=10)

        __Check_1()

        frame1.pack(pady=5)

    
    def _Boundary_Condition_Options_tab4(self, tab):
        frame1 = tk.Frame(tab, bd=2, relief='groove')
        canvas1 = tk.Canvas(frame1)
        canvas2 = tk.Canvas(frame1)
        canvas3 = tk.Canvas(frame1)
        
        def __Check_1():
            if self.var4_1.get():
                for child in canvas1.winfo_children():
                    child.config(state='normal')
                for child in canvas2.winfo_children():
                    child.config(state='normal')
                for child in canvas3.winfo_children():
                    child.config(state='normal')
            else:
                for child in canvas1.winfo_children():
                    child.config(state='disabled')
                for child in canvas2.winfo_children():
                    child.config(state='disabled')
                for child in canvas3.winfo_children():
                    child.config(state='disabled')

        tk.Label(frame1, text='Reservoir Boundary Condition', anchor='w', width=70).pack()
        self.var4_1 = tk.IntVar(frame1, self.WellBC)
        Check1 = tk.Checkbutton(frame1, text='Reservoir BC', variable=self.var4_1, anchor='w', width=70, command=__Check_1)
        Check1.pack()
                
        tk.Label(canvas1, text='Reservoir Type:', anchor='w', width=70).pack()
        self.var4_2 = tk.IntVar(canvas1, self.iWell)
        radio1 = tk.Radiobutton(canvas1, text='Well (rectangular, cylindrical)', variable=self.var4_2, value=1, anchor='w', width=60)
        radio2 = tk.Radiobutton(canvas1, text='Furrow', variable=self.var4_2, value=2, anchor='w', width=60)
        radio3 = tk.Radiobutton(canvas1, text='Wetland', variable=self.var4_2, value=3, anchor='w', width=60)
        radio1.pack()
        radio2.pack()
        radio3.pack()

        tk.Label(canvas2, text='Z-coordinate of the well button', anchor='w', width=55).grid(row=0, column=0)
        tk.Label(canvas2, text='Well radius', anchor='w', width=55).grid(row=1, column=0)
        self.e4_1 = tk.Entry(canvas2, justify='left', width=15)
        self.e4_1.insert(0, self.zWBot)
        self.e4_1.grid(row=0, column=1)
        self.e4_2 = tk.Entry(canvas2, justify='left', width=15)
        self.e4_2.insert(0, self.Radius)
        self.e4_2.grid(row=1, column=1)

        tk.Label(canvas3, text='Pumping (+) or injection (-) rate', anchor='w', width=55).grid(row=0, column=0)
        tk.Label(canvas3, text='Z-coordinate of the initial water level', anchor='w', width=55).grid(row=1, column=0)
        tk.Label(canvas3, text='Solute concentration initially in the reservoir', anchor='w', width=55).grid(row=2, column=0)
        self.e4_3 = tk.Entry(canvas3, justify='left', width=15)
        self.e4_3.insert(0, self.WellPump)
        self.e4_3.grid(row=0, column=1)
        self.e4_4 = tk.Entry(canvas3, justify='left', width=15)
        self.e4_4.insert(0, self.zWInit)
        self.e4_4.grid(row=1, column=1)
        self.e4_5 = tk.Entry(canvas3, justify='left', width=15)
        self.e4_5.insert(0, self.cWell)
        self.e4_5.grid(row=2, column=1)
        
        canvas1.pack(padx=10, pady=5)
        ttk.Separator(frame1, orient='horizontal').pack(fill='x', padx=10, pady=5)
        canvas2.pack(padx=10, pady=5)
        ttk.Separator(frame1, orient='horizontal').pack(fill='x', padx=10, pady=5)
        canvas3.pack(padx=10, pady=5)
        
        __Check_1()
        
        frame1.pack(padx=10, pady=10)
        

if __name__ == '__main__':
    main = HYDRUS2DSIMPLE_GUI()
    main.Main_Menu()